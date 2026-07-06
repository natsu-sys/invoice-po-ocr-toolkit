import os
import sys
import json
import re
import unicodedata
import threading
import time
from datetime import datetime
import customtkinter as ctk
from tkinter import filedialog, messagebox
from copy import copy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.ai.documentintelligence.models import AnalyzeDocumentRequest

import barcode_match as bm
import po_to_csv as poc

# ── Path helper: ใช้ได้ทั้งตอน run .py และตอน pack เป็น .exe ──────────────
def get_base_dir():
    """คืน directory ของ .exe หรือ .py ที่กำลังรันอยู่"""
    if getattr(sys, 'frozen', False):
        # กำลังรันเป็น .exe (PyInstaller)
        return os.path.dirname(sys.executable)
    else:
        # กำลังรันเป็น .py ปกติ
        return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = get_base_dir()

# โหลด .env จากโฟลเดอร์เดียวกับ .exe หรือ .py เสมอ
load_dotenv(os.path.join(BASE_DIR, ".env"))

icon_path = os.path.join(BASE_DIR, "icon.ico")


JP_LABELS = {
    # 発行元 (ผู้ขาย/ผู้ออกเอกสาร)
    "VendorName": "発行元名",
    "VendorAddress": "発行元住所",
    "VendorAddressRecipient": "宛先（部署・担当）",

    # 取引先 (ลูกค้า/คู่ค้า)
    "CustomerName": "取引先名",
    "CustomerAddress": "取引先住所",

    # 請求情報
    "InvoiceId": "伝票番号",
    "InvoiceDate": "伝票日付",
    "InvoiceTotal":"請求合計",
    "DueDate": "支払期日",
    "PurchaseOrder": "注文番号",
    "PaymentTerms": "支払条件",
    "CurrencyCode": "通貨",

    # 金額
    "SubTotal": "小計",
    "Tax": "消費税",
    "Total": "合計",

    # その他 (ถ้ามี)
    "BillingAddress": "請求先住所",
    "ShippingAddress": "納品先住所",
    "Page": "ページ",
}

# ======================
# CONFIG
# ======================
ENDPOINT = os.getenv("AZURE_DOC_ENDPOINT")
KEY = os.getenv("AZURE_DOC_KEY")


def _app_base_dir():
    """โฟลเดอร์ที่โปรแกรมอยู่ — รองรับทั้งรันเป็น .py และแพ็คเป็น .exe (PyInstaller)"""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


APP_DIR = _app_base_dir()
SETTINGS_PATH = os.path.join(APP_DIR, "settings.json")


def load_settings():
    try:
        with open(SETTINGS_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_settings(**kv):
    """อัปเดต settings.json (merge) — เงียบถ้าเขียนไม่ได้"""
    try:
        data = load_settings()
        data.update(kv)
        with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _default_out_dir():
    """default = โฟลเดอร์ output ข้างโปรแกรม (พกพาได้); ถ้าเขียนไม่ได้ fallback ไป Documents"""
    beside = os.path.join(APP_DIR, "output")
    try:
        os.makedirs(beside, exist_ok=True)
        return beside
    except Exception:
        fb = os.path.join(os.path.expanduser("~"), "Documents", "発注書CSV")
        os.makedirs(fb, exist_ok=True)
        return fb


# จำที่เลือกล่าสุด ถ้ามี (และยังมีอยู่จริง) ไม่งั้นใช้ default ข้างโปรแกรม
DEFAULT_OUT_DIR = load_settings().get("out_dir") or _default_out_dir()

MODEL_ID = "prebuilt-invoice"

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ======================
# ENHANCED FIELD MAPPING
# ======================
# Field mapping: Standard name → [English variants, Japanese variants, Custom variants]
ENHANCED_FIELD_MAPPING = {
    # ========================================
    # INVOICE IDENTIFIERS
    # ========================================
    "InvoiceId": [
        # English
        "InvoiceId", "Invoice No", "Invoice Number", "Invoice #", "Doc Number",
        # Japanese
        "請求書番号", "請求No", "伝票番号", "請求番号", "No.",
        # Custom
        "Bill No", "Billing Number"
    ],

    # ========================================
    # DATES
    # ========================================
    "InvoiceDate": [
        "InvoiceDate", "Invoice Date", "Date", "Issue Date", "Issued Date", "Document Date",
        "請求日", "発行日", "日付", "作成日", "発行年月日"
    ],

    "DueDate": [
        "DueDate", "Due Date", "Payment Due", "Payment Date", "Maturity Date",
        "支払期限", "お支払期限", "期日", "支払日", "納期"
    ],

    "ServiceDate": [
        "Service Date", "Delivery Date", "Ship Date",
        "納品日", "配送日", "出荷日"
    ],

    # ========================================
    # VENDOR (SELLER) INFORMATION
    # ========================================
    "VendorName": [
        "VendorName", "Vendor", "From", "Supplier", "Seller", "Provider",
        "発行者", "会社名", "発行元", "売主", "販売者", "業者名"
    ],

    "VendorAddress": [
        "VendorAddress", "Vendor Address", "From Address", "Seller Address",
        "住所", "発行者住所", "所在地", "本社所在地"
    ],

    "VendorTaxId": [
        "VendorTaxId", "Tax ID", "VAT Number", "Tax Number",
        "法人番号", "税務番号", "事業者番号"
    ],

    # Japanese-specific vendor fields
    "VendorRegistrationNumber": [
        "Registration Number", "Qualified Invoice Number",
        "適格請求書発行事業者登録番号", "登録番号", "インボイス登録番号"
    ],

    "VendorCorporateNumber": [
        "Corporate Number", "Company Number",
        "法人番号", "会社番号"
    ],

    # ========================================
    # CUSTOMER (BUYER) INFORMATION
    # ========================================
    "CustomerName": [
        "CustomerName", "Customer", "Bill To", "To", "Buyer", "Client",
        "宛先", "御中", "様", "請求先", "お客様", "顧客名", "請求先名"
    ],

    "CustomerAddress": [
        "CustomerAddress", "Customer Address", "Bill To Address", "Buyer Address",
        "宛先住所", "お届け先", "顧客住所"
    ],

    "CustomerId": [
        "Customer ID", "Customer Code", "Client ID", "Account Number",
        "顧客番号", "得意先コード", "取引先コード"
    ],

    "CustomerTaxId": [
        "Customer Tax ID", "Customer VAT",
        "顧客税務番号"
    ],

    # ========================================
    # AMOUNTS (MONEY)
    # ========================================
    "SubTotal": [
        "SubTotal", "Subtotal", "Net Amount", "Amount Before Tax", "Gross Amount",
        "小計", "税抜金額", "合計(税抜)", "税抜合計", "本体価格"
    ],

    "TotalTax": [
        "TotalTax", "Tax", "VAT", "Sales Tax", "Tax Amount", "GST",
        "消費税", "税額", "消費税額", "税金", "消費税等"
    ],

    "InvoiceTotal": [
        "InvoiceTotal", "Total", "Grand Total", "Amount Due", "Total Amount", "Final Amount",
        "合計", "総額", "請求金額", "お支払金額", "合計金額", "御請求金額", "請求額"
    ],

    "AmountDue": [
        "Amount Due", "Balance Due", "Outstanding Amount",
        "未払金額", "残高", "支払残高"
    ],

    "Discount": [
        "Discount", "Discount Amount", "Rebate",
        "割引", "値引", "値引額", "ディスカウント"
    ],

    "ShippingCost": [
        "Shipping", "Shipping Cost", "Delivery Fee", "Freight",
        "送料", "配送料", "運賃", "配送費"
    ],

    # ========================================
    # PAYMENT INFORMATION
    # ========================================
    "PaymentTerm": [
        "Payment Term", "Payment Terms", "Terms", "Payment Conditions",
        "支払条件", "支払期間", "決済条件"
    ],

    "BankName": [
        "Bank Name", "Bank", "Financial Institution",
        "銀行名", "金融機関", "金融機関名"
    ],

    "AccountNumber": [
        "Account Number", "Account No", "Bank Account",
        "口座番号", "アカウント番号", "預金口座番号"
    ],

    "AccountHolder": [
        "Account Holder", "Account Name", "Beneficiary",
        "口座名義", "名義人", "受取人名"
    ],

    "BranchCode": [
        "Branch Code", "Branch", "Sort Code",
        "支店コード", "店番", "支店番号"
    ],

    "SwiftCode": [
        "SWIFT Code", "BIC", "SWIFT/BIC",
        "スウィフトコード", "BICコード"
    ],

    # ========================================
    # REFERENCE NUMBERS
    # ========================================
    "PurchaseOrder": [
        "Purchase Order", "PO Number", "PO", "Order Number",
        "発注番号", "注文番号", "PO番号", "購買番号"
    ],

    "QuotationNumber": [
        "Quotation Number", "Quote Number", "Estimate Number",
        "見積番号", "見積No", "見積書番号"
    ],

    "ContractNumber": [
        "Contract Number", "Contract No", "Agreement Number",
        "契約番号", "契約No", "契約書番号"
    ],

    "ProjectCode": [
        "Project Code", "Project Number", "Project ID", "Job Number",
        "プロジェクトコード", "案件番号", "プロジェクト番号", "工事番号"
    ],

    "ReferenceNumber": [
        "Reference Number", "Ref No", "Reference",
        "参照番号", "参考番号", "リファレンス番号"
    ],

    # ========================================
    # BUSINESS/ACCOUNTING
    # ========================================
    "CostCenter": [
        "Cost Center", "CC", "Profit Center",
        "コストセンター", "原価部門", "経費部門"
    ],

    "Department": [
        "Department", "Dept", "Division",
        "部署", "部門", "課"
    ],

    "GLCode": [
        "GL Code", "GL Account", "Account Code", "Ledger Code",
        "勘定科目コード", "GL科目", "会計コード"
    ],

    # ========================================
    # TAX DETAILS (JAPANESE SPECIFIC)
    # ========================================
    "TaxRate10": [
        "Standard Tax Rate", "10% Tax",
        "標準税率", "10%税", "消費税10%"
    ],

    "TaxRate8": [
        "Reduced Tax Rate", "8% Tax",
        "軽減税率", "8%税", "消費税8%"
    ],

    "TaxCategory": [
        "Tax Category", "Tax Type", "Tax Classification",
        "税区分", "税種別", "課税区分"
    ],

    "TaxableAmount": [
        "Taxable Amount", "Tax Base",
        "課税額", "課税対象額", "税抜金額"
    ],

    # ========================================
    # SHIPPING/DELIVERY
    # ========================================
    "ShippingAddress": [
        "Shipping Address", "Delivery Address", "Ship To",
        "配送先住所", "納品先住所", "送付先"
    ],

    "DeliveryDate": [
        "Delivery Date", "Ship Date", "Dispatch Date",
        "配送日", "納品日", "出荷日"
    ],

    "TrackingNumber": [
        "Tracking Number", "Tracking No", "Waybill Number",
        "追跡番号", "荷物番号", "送り状番号"
    ],

    # ========================================
    # CURRENCY
    # ========================================
    "Currency": [
        "Currency", "Currency Code",
        "通貨", "通貨コード", "貨幣"
    ],

    # ========================================
    # LINE ITEM FIELDS
    # ========================================
    "ItemNo": [
        "Item No", "Line", "No", "#", "Line No",
        "No.", "番号", "項番", "行番号"
    ],

    "Description": [
        "Description", "Item", "Product", "Service", "Item Description",
        "品名", "品目", "商品名", "摘要", "内容", "明細"
    ],

    "ProductCode": [
        "Product Code", "Code", "SKU", "Item Code", "Part Number",
        "商品コード", "品番", "型番", "品目コード", "製品番号"
    ],

    "Quantity": [
        "Quantity", "Qty", "Quantity Ordered", "Units", "Amount",
        "数量", "個数", "数", "本数", "個"
    ],

    "Unit": [
        "Unit", "Unit of Measure", "UOM", "Unit Type",
        "単位", "単", "計量単位"
    ],

    "UnitPrice": [
        "Unit Price", "Price", "Rate", "Unit Cost",
        "単価", "価格", "単位価格", "単位金額"
    ],

    "Amount": [
        "Amount", "Line Total", "Total", "Line Amount", "Extended Price",
        "金額", "小計", "合計額", "明細金額"
    ],

    "TaxRate": [
        "Tax Rate", "VAT Rate", "Tax %",
        "税率", "消費税率", "税％"
    ],

    "TaxAmount": [
        "Tax Amount", "Tax", "VAT Amount",
        "税額", "消費税額", "税金額"
    ],

    # ========================================
    # OTHER FIELDS
    # ========================================
    "Notes": [
        "Notes", "Remarks", "Comments", "Memo",
        "備考", "注記", "メモ", "特記事項"
    ],

    "Terms": [
        "Terms", "Terms and Conditions", "T&C",
        "条件", "約款", "規約"
    ],
}



# ======================
# Progress Simulator
# ======================
class ProgressSimulator:
    """Simulate smooth progress while waiting for API"""

    def __init__(self, callback, start=0.0, target=0.9, duration=3.0):
        self.callback = callback
        self.start = start
        self.target = target
        self.duration = duration
        self.stop_flag = False
        self.thread = None

    def start_simulation(self):
        """เริ่ม simulate progress"""
        self.stop_flag = False
        self.thread = threading.Thread(target=self._simulate, daemon=True)
        self.thread.start()

    def stop_simulation(self):
        """หยุด simulation"""
        self.stop_flag = True
        if self.thread:
            self.thread.join(timeout=0.5)

    def _simulate(self):
        """Simulate smooth progress"""
        steps = 50  # 50 steps
        interval = self.duration / steps
        increment = (self.target - self.start) / steps

        current = self.start
        for _ in range(steps):
            if self.stop_flag:
                break

            current += increment
            self.callback(min(current, self.target))
            time.sleep(interval)


# ======================
# Japanese Helper
# ======================
class JapaneseHelper:
    """日本語処理ヘルパー"""

    def __init__(self):
        self.date_patterns = [
            (r'令和(\d+)年(\d+)月(\d+)日', self._reiwa_to_date),
            (r'平成(\d+)年(\d+)月(\d+)日', self._heisei_to_date),
            (r'(\d{4})年(\d{1,2})月(\d{1,2})日', self._standard_jp_date),
            (r'(\d{4})/(\d{1,2})/(\d{1,2})', self._slash_date),
        ]

    def normalize_text(self, text: str) -> str:
        if not text:
            return ""
        text = unicodedata.normalize('NFKC', text)
        # Azure selection marks: :selected:=ติ๊กแล้ว→☑, :unselected:=ไม่ติ๊ก→ลบ (มักเป็น noise)
        text = text.replace(":selected:", " ☑ ").replace(":unselected:", " ")
        text = re.sub(r'[\s\u3000]+', ' ', text)
        return text.strip()

    def normalize_keep_newline(self, text: str) -> str:
        """normalize แต่เก็บ newline ไว้ สำหรับ cell ที่มีหลายค่า stacked"""
        if not text:
            return ""
        text = unicodedata.normalize('NFKC', text)
        text = text.replace(":selected:", " \u2611 ").replace(":unselected:", " ")
        text = re.sub(r'[ \t\u3000]+', ' ', text)   # normalize spaces only
        text = re.sub(r'\n{2,}', '\n', text)          # reduce multi-newline
        return text.strip()

    def parse_date(self, date_str: str) -> str:
        if not date_str:
            return ""
        date_str = self.normalize_text(date_str)
        for pattern, converter in self.date_patterns:
            match = re.search(pattern, date_str)
            if match:
                try:
                    return converter(match)
                except:
                    continue
        return date_str

    def _reiwa_to_date(self, match) -> str:
        year = int(match.group(1)) + 2018
        month = int(match.group(2))
        day = int(match.group(3))
        return f"{year}-{month:02d}-{day:02d}"

    def _heisei_to_date(self, match) -> str:
        year = int(match.group(1)) + 1988
        month = int(match.group(2))
        day = int(match.group(3))
        return f"{year}-{month:02d}-{day:02d}"

    def _standard_jp_date(self, match) -> str:
        year = match.group(1)
        month = int(match.group(2))
        day = int(match.group(3))
        return f"{year}-{month:02d}-{day:02d}"

    def _slash_date(self, match) -> str:
        year = match.group(1)
        month = int(match.group(2))
        day = int(match.group(3))
        return f"{year}-{month:02d}-{day:02d}"


# ======================
# Smart Field Mapper
# ======================
class SmartFieldMapper:
    """Map field names to standard names using enhanced mapping"""

    def __init__(self):
        self.mapping = ENHANCED_FIELD_MAPPING
        self.unknown_fields = {}  # Track unmapped fields

    def map_field_name(self, original_name: str) -> str:
        """
        Map original field name to standard name

        Args:
            original_name: Original field name from Azure

        Returns:
            Standard field name (or original if no mapping found)
        """
        if not original_name:
            return original_name

        # 1. Try exact match (case-sensitive)
        for standard_name, variants in self.mapping.items():
            if original_name in variants:
                return standard_name

        # 2. Try exact match (case-insensitive)
        original_lower = original_name.lower()
        for standard_name, variants in self.mapping.items():
            for variant in variants:
                if original_lower == variant.lower():
                    return standard_name

        # 3. Try partial match (contains)
        for standard_name, variants in self.mapping.items():
            for variant in variants:
                # Check if variant is in original name
                if variant.lower() in original_lower:
                    return standard_name
                # Check if original name is in variant
                if original_lower in variant.lower():
                    return standard_name

        # 4. No mapping found - track as unknown
        self.unknown_fields[original_name] = True

        return original_name  # Return as-is

    def get_unknown_fields(self) -> list:
        """Get list of fields that couldn't be mapped"""
        return sorted(self.unknown_fields.keys())

# ======================
# Field Extraction
# ======================
def field_to_text(f):
    if f is None:
        return ""
    c = getattr(f, "content", None)
    if c:
        return str(c).strip()
    for attr in ("value_string", "value_phone_number", "value_date", "value_time"):
        v = getattr(f, attr, None)
        if v is not None:
            return str(v)
    for attr in ("value_number", "value_integer"):
        v = getattr(f, attr, None)
        if v is not None:
            return str(v)
    vc = getattr(f, "value_currency", None)
    if vc is not None:
        amt = getattr(vc, "amount", "")
        return str(amt)
    return ""


# ======================
# Azure Analysis
# ======================
def analyze_invoice(pdf_path: str, log_fn=None):
    client = DocumentIntelligenceClient(
        endpoint=ENDPOINT,
        credential=AzureKeyCredential(KEY)
    )
    with open(pdf_path, "rb") as f:
        doc_bytes = f.read()

    def _run():
        poller = client.begin_analyze_document(
            model_id=MODEL_ID,
            body=AnalyzeDocumentRequest(bytes_source=doc_bytes),
        )
        return poller.result()

    return bm.call_with_retry(_run, log_fn=log_fn)

def analyze_document(pdf_path: str, log_fn=None):
    """ใช้ prebuilt-layout สำหรับเอกสารทั่วไป (ไม่ใช่ invoice)"""
    client = DocumentIntelligenceClient(
        endpoint=ENDPOINT,
        credential=AzureKeyCredential(KEY)
    )
    with open(pdf_path, "rb") as f:
        doc_bytes = f.read()

    def _run():
        poller = client.begin_analyze_document(
            model_id="prebuilt-layout",
            body=AnalyzeDocumentRequest(bytes_source=doc_bytes),
        )
        return poller.result()

    return bm.call_with_retry(_run, log_fn=log_fn)




def extract_fields(result, jp_helper):
    """Extract fields with COMPLETE table support"""
    print("extract_fields called, jp_helper =", type(jp_helper))
    docs = getattr(result, "documents", None) or []
    field_mapper = SmartFieldMapper()

    # ========================================
    # EXTRACT HEADER FIELDS (เดิม)
    # ========================================
    header = {}
    header_original = {}

    if docs:
        fields = docs[0].fields or {}

        for key, field in fields.items():
            if key == "Items":
                continue

            value = field_to_text(field)

            # Japanese date processing
            if any(x in key for x in ['Date', '日']):
                value = jp_helper.parse_date(value)
            else:
                value = jp_helper.normalize_text(value)

            # Map to standard name
            standard_name = field_mapper.map_field_name(key)

            # Store both
            header_original[key] = value
            header[standard_name] = value

    # ========================================
    # EXTRACT ITEMS - PRIORITIZE TABLES!
    # ========================================
    items = []

    # ✅ NEW: Extract from TABLES first (like Azure Studio)
    tables = getattr(result, "tables", None) or []

    if tables:
        print(f"\n=== Found {len(tables)} table(s) ===")

        for table_idx, table in enumerate(tables, 1):
            print(f"Processing Table {table_idx}: {table.row_count} rows x {table.column_count} cols")

            # Build table data structure
            table_data = {}
            for cell in table.cells:
                row_idx = cell.row_index
                col_idx = cell.column_index

                if row_idx not in table_data:
                    table_data[row_idx] = {}

                table_data[row_idx][col_idx] = cell.content

            # Find header row (usually row 0 or 1)
            header_row_idx = None
            headers = []

            for row_idx in range(min(3, len(table_data))):
                if row_idx in table_data:
                    row_values = list(table_data[row_idx].values())
                    row_text = ' '.join(str(v) for v in row_values).lower()

                    # Check if contains header keywords
                    if any(keyword in row_text for keyword in [
                        '品名', 'description', '数量', 'quantity', '金額', 'amount',
                        '単価', 'price', '税', 'tax', '合計', 'total'
                    ]):
                        header_row_idx = row_idx
                        headers = table_data[row_idx]
                        break

            # If no header found, use first row as header
            if header_row_idx is None and 0 in table_data:
                header_row_idx = 0
                headers = table_data[0]

            # Extract data rows
            data_start_row = header_row_idx + 1 if header_row_idx is not None else 0

            for row_idx in range(data_start_row, len(table_data)):
                if row_idx not in table_data:
                    continue

                row_data = table_data[row_idx]
                item_data = {}

                # Map each column to field
                for col_idx, value in row_data.items():
                    # Get header name for this column
                    header_name = headers.get(col_idx, f"Column_{col_idx}")

                    # Normalize value
                    value_normalized = jp_helper.normalize_text(str(value))

                    # Skip empty values
                    if not value_normalized or value_normalized in ['', '-', '—', 'None']:
                        continue

                    # Map to standard field name
                    standard_name = field_mapper.map_field_name(str(header_name))

                    item_data[standard_name] = value_normalized

                # Only add if has meaningful data
                if item_data and (
                        item_data.get('Description') or
                        item_data.get('Amount') or
                        item_data.get('Quantity')
                ):
                    # Add metadata
                    item_data['_source'] = f'Table_{table_idx}'
                    item_data['_row'] = row_idx

                    items.append(item_data)

            print(
                f"  → Extracted {len([i for i in items if i.get('_source') == f'Table_{table_idx}'])} items from Table {table_idx}")

    # ========================================
    # FALLBACK: Extract from Documents.Items (ถ้า tables ไม่มี)
    # ========================================
    if not items and docs:
        print("No tables found, extracting from Documents.Items...")

        fields = docs[0].fields or {}
        items_field = fields.get("Items")

        if items_field and getattr(items_field, "value_array", None):
            for idx, item in enumerate(items_field.value_array):
                obj = item.value_object or {}
                item_data = {}

                for key, field in obj.items():
                    value = field_to_text(field)
                    value = jp_helper.normalize_text(value)
                    standard_name = field_mapper.map_field_name(key)
                    item_data[standard_name] = value

                if item_data:
                    item_data['_source'] = 'Documents'
                    items.append(item_data)

    # ========================================
    # CLEANUP & RETURN
    # ========================================
    # Remove metadata fields from display
    for item in items:
        if '_row' in item:
            del item['_row']
        if '_source' in item:
            del item['_source']

    # Add unknown fields tracking
    unknown_fields = field_mapper.get_unknown_fields()
    if unknown_fields:
        header['_unknown_fields'] = ', '.join(unknown_fields)
        header['_original_fields'] = header_original

    print(f"=== Total extracted: {len(items)} items ===\n")

    return header, items

# ======================
# Beautiful Excel Export (abbreviated for space)
# ======================

# ======================
# Table Rendering (Azure -> Excel)
# ======================
from openpyxl.cell.cell import MergedCell

def pick_best_items_table(result, jp_helper):
    """
    Pick the table that most likely represents the main line-items grid.
    Returns (table_index_1based, table_obj) or (None, None) if no tables.
    """
    tables = getattr(result, "tables", None) or []
    if not tables:
        return None, None

    def score_headers(headers: dict) -> float:
        text = " ".join(str(v) for v in headers.values()).lower()
        keywords = ["伝票", "出荷", "納期", "品名", "品番", "数量", "単価", "金額", "重量", "員数", "受注", "訂正", "備考"]
        return sum(1 for k in keywords if k in text)

    best = None  # (score, idx, table)
    for idx, table in enumerate(tables, start=1):
        # Build quick header row guess (top 5 rows)
        rows = {}
        for cell in table.cells:
            rows.setdefault(cell.row_index, {})
            rows[cell.row_index][cell.column_index] = jp_helper.normalize_text(cell.content or "")

        header_row = None
        header = None
        for r in sorted(rows.keys())[:5]:
            row_text = " ".join(str(v) for v in rows[r].values()).lower()
            if any(k in row_text for k in ["品名", "数量", "単価", "金額", "出荷", "伝票"]):
                header_row = r
                header = rows[r]
                break
        if header is None and 0 in rows:
            header_row = 0
            header = rows[0]
        if not header:
            continue

        s = score_headers(header)
        # more rows -> more likely to be main table
        s += min(table.row_count, 40) / 20.0
        if best is None or s > best[0]:
            best = (s, idx, table)

    if not best:
        return None, None
    return best[1], best[2]


def _safe_set_value(ws, r, c, value):
    """Write value into a worksheet cell, redirecting to top-left if cell is a merged-cell placeholder."""
    cell = ws.cell(r, c)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                cell = ws.cell(mr.min_row, mr.min_col)
                break
    cell.value = value
    return cell


def render_azure_table(ws, table, start_row=1, start_col=1, jp_helper=None):
    """
    Render Azure table into worksheet using row/col indices and spans (merge).
    แปลงวันที่และตัวเลขให้เป็น Excel type ที่ใช้ filter/SUM ได้
    """
    from datetime import datetime as _dt
    from openpyxl.styles import numbers as xl_numbers
    # Write content and merge spans
    for tcell in table.cells:
        r = start_row + tcell.row_index
        c = start_col + tcell.column_index

        # แปลงค่าก่อนเขียน
        raw = tcell.content or ""
        val = smart_cell_value(raw, jp_helper)
        cell = _safe_set_value(ws, r, c, val)
        if cell is not None and isinstance(val, _dt):
            cell.number_format = "YYYY/MM/DD"
        elif cell is not None and isinstance(val, (int, float)):
            cell.number_format = "#,##0"

        # rs = getattr(tcell, "row_span", 1) or 1
        # cs = getattr(tcell, "column_span", 1) or 1
        # if rs > 1 or cs > 1:
        #     ws.merge_cells(start_row=r, start_column=c, end_row=r + rs - 1, end_column=c + cs - 1)


def format_table_sheet(ws, header_row=1):
    """
    Make the rendered table look closer to the original: borders, wrap, auto-width,
    numeric alignment, header styling.
    """
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    base_font = Font(name="MS Gothic", size=10)
    header_font = Font(name="MS Gothic", size=10, bold=True, color="FFFFFF")

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < header_row or max_col < 1:
        return

    # base style
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = base_font

    # header style
    header_fill = PatternFill("solid", fgColor="366092")
    header_font = Font(name="MS Gothic", size=10, bold=True, color="FFFFFF")
    for c in range(1, max_col + 1):
        cell = ws.cell(header_row, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 22

    # auto width
    for c in range(1, max_col + 1):
        cell = ws.cell(header_row, c)
        cell.fill = header_fill
        cell.font = header_font  # ✅ บังคับหัวตาราง
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        max_len = 0
        for r in range(1, max_row + 1):
            v = ws.cell(r, c).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(8, max_len * 1.2), 55)

    # align by header keywords
    headers = [str(ws.cell(header_row, c).value or "").replace("\n", "").strip() for c in range(1, max_col + 1)]
    for c, h in enumerate(headers, start=1):
        if any(k in h for k in ["単価", "金額", "重量", "員数", "数量"]):
            for r in range(header_row + 1, max_row + 1):
                ws.cell(r, c).alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        if any(k in h for k in ["出荷日", "納期", "発行日"]):
            for r in range(header_row + 1, max_row + 1):
                ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        if any(k in h for k in ["品名", "摘要", "品名寸法"]):
            ws.column_dimensions[get_column_letter(c)].width = max(ws.column_dimensions[get_column_letter(c)].width, 35)
            for r in range(header_row + 1, max_row + 1):
                ws.cell(r, c).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"


def create_layout_sheet(wb, result, jp_helper, log_fn):
    """
    Sheet 1: 明細レイアウト
    Flattens Azure row_span/col_span properly so every data cell appears in Excel.
    Strategy: expand spanned cells into a flat grid first, then write to Excel.
    """
    ws = wb.create_sheet("明細レイアウト")

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(name="MS Gothic", size=10, bold=True, color="FFFFFF")
    base_font = Font(name="MS Gothic", size=10)
    subtotal_font = Font(name="MS Gothic", size=10, bold=True)
    subtotal_fill = PatternFill("solid", fgColor="EBF3FB")

    best_idx, best_table = pick_best_items_table(result, jp_helper)
    if not best_table:
        ws["A1"] = "明細テーブルが検出されませんでした。"
        log_fn("⚠ No layout table detected.\n")
        return

    n_rows = best_table.row_count
    n_cols = best_table.column_count

    # ── Step 1: Build flat grid (expand spans) ───────────────────────────────
    # grid[r][c] = text value — every cell gets a value (span cells repeat)
    grid = [[" " for _ in range(n_cols)] for _ in range(n_rows)]

    for tcell in best_table.cells:
        r, c = tcell.row_index, tcell.column_index
        rs = max(getattr(tcell, "row_span", 1) or 1, 1)
        cs = max(getattr(tcell, "column_span", 1) or 1, 1)
        val = jp_helper.normalize_text(tcell.content or "")

        # Fill every cell covered by this span with the value
        for dr in range(rs):
            for dc in range(cs):
                rr, cc = r + dr, c + dc
                if rr < n_rows and cc < n_cols:
                    # Only write primary value once; repeated span cells get ""
                    grid[rr][cc] = val if (dr == 0 and dc == 0) else ""

    # ── Step 2: Find header row ───────────────────────────────────────────────
    hdr_row_idx = 0
    for r in range(min(5, n_rows)):
        row_text = " ".join(grid[r])
        if any(k in row_text for k in ["品名", "数量", "単価", "金額", "出荷", "伝票", "品番", "日付"]):
            hdr_row_idx = r
            break

    col_headers = grid[hdr_row_idx]  # list of header names per column

    # ── Step 3: Write flat grid to Excel ─────────────────────────────────────
    right_align_kw = ["単価", "金額", "重量", "員数", "数量", "合計", "小計"]
    center_align_kw = ["出荷日", "納期", "日付", "No", "番号"]
    wide_col_kw = ["品名", "摘要", "品名寸法", "商品名"]
    subtotal_kw = ["小計", "合計", "総計", "頁計"]

    for r in range(n_rows):
        excel_row = r + 1
        row_vals = grid[r]
        is_header = (r == hdr_row_idx)
        row_text_all = " ".join(row_vals)
        is_subtotal = (not is_header) and any(k in row_text_all for k in subtotal_kw)

        for c in range(n_cols):
            raw_val = row_vals[c]
            excel_col = c + 1
            col_name = col_headers[c] if c < len(col_headers) else ""

            # แปลงวันที่/ตัวเลขเฉพาะ data rows (ไม่แปลง header)
            from datetime import datetime as _dt
            if not is_header and not is_subtotal:
                val = smart_cell_value(raw_val, jp_helper)
            else:
                val = raw_val

            cell = ws.cell(excel_row, excel_col, val)

            if is_header:
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.row_dimensions[excel_row].height = 24
            elif is_subtotal:
                cell.font = subtotal_font
                cell.fill = subtotal_fill
                if any(k in col_name for k in right_align_kw):
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                ws.row_dimensions[excel_row].height = 18
            else:
                cell.font = base_font
                if isinstance(val, _dt):
                    cell.number_format = "YYYY/MM/DD"
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                elif isinstance(val, (int, float)):
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
                elif any(k in col_name for k in right_align_kw):
                    cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
                elif any(k in col_name for k in center_align_kw):
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                ws.row_dimensions[excel_row].height = 18

            cell.border = border

    # ── Pass 2: Clear fill on spanned rows to prevent color bleed, then merge ──
    no_fill = PatternFill(fill_type=None)
    for tcell in best_table.cells:
        r, c = tcell.row_index, tcell.column_index
        rs = max(getattr(tcell, "row_span", 1) or 1, 1)
        cs = max(getattr(tcell, "column_span", 1) or 1, 1)
        if rs > 1 or cs > 1:
            er = min(r + rs - 1, n_rows - 1)
            ec = min(c + cs - 1, n_cols - 1)
            # Clear fill on all spanned cells (except top-left) to remove color bleed
            for dr in range(rs):
                for dc in range(cs):
                    if dr == 0 and dc == 0:
                        continue
                    rr, cc = r + dr, c + dc
                    if rr < n_rows and cc < n_cols:
                        try:
                            ws.cell(rr + 1, cc + 1).fill = no_fill
                        except Exception:
                            pass
            try:
                ws.merge_cells(
                    start_row=r + 1, start_column=c + 1,
                    end_row=er + 1, end_column=ec + 1
                )
            except Exception:
                pass

    # ── Step 4: Auto column widths ────────────────────────────────────────────
    for c in range(n_cols):
        excel_col = c + 1
        col_name = col_headers[c] if c < len(col_headers) else ""
        max_len = 0
        for r in range(n_rows):
            v = grid[r][c]
            if v and v.strip():
                length = sum(2 if ord(ch) > 127 else 1 for ch in str(v))
                max_len = max(max_len, length)
        if any(k in col_name for k in wide_col_kw):
            ws.column_dimensions[get_column_letter(excel_col)].width = max(30, min(max_len * 0.9, 60))
        else:
            ws.column_dimensions[get_column_letter(excel_col)].width = max(8, min(max_len * 0.9, 30))

    ws.freeze_panes = "A2"
    log_fn(f"✓ Layout sheet: Table_{best_idx} — {n_rows} rows × {n_cols} cols\n")


def create_beautiful_excel(header: dict, pdf_path: str, out_xlsx: str, log_fn, result):
    """
    Output Excel with 2 sheets:
      Sheet 1: 明細レイアウト — line-items table rendered like original invoice
      Sheet 2: 請求書情報     — structured header fields (existing behavior)
    """
    wb = Workbook()
    wb.remove(wb.active)

    jp_helper = JapaneseHelper()

    # ── Sheet 1: Structured data (請求書情報) ────────────────────
    title_font = Font(name='MS Gothic', size=14, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    label_font = Font(name='MS Gothic', size=11, bold=True)
    label_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    value_font = Font(name='MS Gothic', size=11)
    border_thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws_header = wb.create_sheet("請求書情報")
    ws_header.merge_cells('A1:D1')
    ws_header['A1'] = 'INVOICE DETAILS'
    ws_header['A1'].font = title_font
    ws_header['A1'].fill = title_fill
    ws_header['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_header.row_dimensions[1].height = 25

    row = 3
    ws_header.cell(row, 1, 'File Name:').font = label_font
    ws_header.cell(row, 2, os.path.basename(pdf_path)).font = value_font
    row += 1
    ws_header.cell(row, 1, 'Processed:').font = label_font
    ws_header.cell(row, 2, datetime.now().strftime("%Y-%m-%d %H:%M:%S")).font = value_font
    row += 2

    priority_fields = [
        'InvoiceId', 'InvoiceDate', 'DueDate',
        'VendorName', 'VendorAddress', 'VendorTaxId',
        'CustomerName', 'CustomerAddress',
        'SubTotal', 'TotalTax', 'InvoiceTotal'
    ]

    for field_name in priority_fields:
        val = header.get(field_name)
        if val is None or val == "":
            continue
        label = JP_LABELS.get(field_name, field_name)
        ws_header.cell(row, 1, label).font = label_font
        ws_header.cell(row, 1).fill = label_fill
        ws_header.cell(row, 1).border = border_thin
        ws_header.cell(row, 2, str(val)).font = value_font
        ws_header.cell(row, 2).border = border_thin
        row += 1

    for key in header.keys():
        if key in priority_fields:
            continue
        if str(key).startswith("_"):
            continue
        val = header.get(key)
        if val is None or val == "":
            continue
        if isinstance(val, (dict, list, tuple, set)):
            val = json.dumps(val, ensure_ascii=False)
        label = JP_LABELS.get(str(key), str(key))
        ws_header.cell(row, 1, label).font = label_font
        ws_header.cell(row, 1).fill = label_fill
        ws_header.cell(row, 1).border = border_thin
        ws_header.cell(row, 2, str(val)).font = value_font
        ws_header.cell(row, 2).border = border_thin
        row += 1

    ws_header.column_dimensions['A'].width = 25
    ws_header.column_dimensions['B'].width = 50

    # ── Sheet 2: Layout (明細レイアウト) ─────────────────────────
    create_layout_sheet(wb, result, jp_helper, log_fn)

    wb.save(out_xlsx)
    log_fn(f"✓ {os.path.basename(out_xlsx)}\n")


# ======================
# Export Many to ONE Excel
# ======================

# ======================

def normalize_part_no(part_code: str) -> str:
    """
    Normalize Part No.:
    1. OCR correction: I/T → 1 ในทุก segment ที่มีตัวเลขปนอยู่
    2. ตัด trailing -0 ออก
    """
    import re
    code = part_code.strip().upper()

    def fix_segment(seg):
        if re.search(r'\d', seg):
            seg = seg.replace('I', '1').replace('T', '1')
        return seg

    parts = code.split('-')
    parts = [fix_segment(p) for p in parts]
    code = '-'.join(parts)
    # ตัด trailing segment ที่เป็นศูนย์ล้วนทุกรูปแบบ: -0, -00, -000
    code = re.sub(r'(-0+)+$', '', code)
    return code


def extract_part_no(text: str) -> str:
    """
    ดึง Part No. จากข้อความ เช่น "50010-93652-0 シャフト" → "50010-93652-0"
    รองรับ: ขึ้นต้นด้วยตัวอักษรหรือตัวเลข, segment ที่เป็นตัวอักษรล้วน (เช่น LS2-C-7-01)
    เลือก candidate ที่มีตัวเลขและยาวที่สุด
    """
    import re
    pattern = re.compile(r'\b([A-Z0-9]{2,}(?:-[A-Z0-9]+){1,})\b')
    text_up = text.upper()
    candidates = pattern.findall(text_up)
    valid = [c for c in candidates if re.search(r'\d', c) and len(c) >= 6]
    if not valid:
        return ""
    return normalize_part_no(max(valid, key=len))


def detect_meisai_file(pdf_path: str) -> bool:
    """
    ตรวจสอบว่าไฟล์นี้คือ 注文明細票 ไหม
    โดยดูจากชื่อไฟล์หรือ pattern ของ barcode column (XXXXXXXX-XXX)
    """
    import re
    fname = os.path.basename(pdf_path).lower()
    # ชื่อไฟล์ไม่มี _0001, _0002 suffix = น่าจะเป็น 注文明細票
    if re.search(r'_\d{4}\.pdf$', fname):
        return False
    return True  # ไฟล์ที่ไม่มี suffix ถือว่าเป็น 注文明細票


def build_barcode_map(meisai_paths: list, log_fn, progress_callback=None, base_progress=0.0, progress_range=0.2) -> dict:
    """
    อ่านไฟล์ 注文明細票 แล้วสร้าง dict {part_no: [barcode1, barcode2, ...]}
    ถ้า Part No. ซ้ำกัน เก็บ barcode เป็น list ตามลำดับที่ปรากฏ
    ตอน lookup จะ pop ออกทีละอันตามลำดับ (FIFO)
    """
    import re
    barcode_map = {}  # {part_no: [barcode, ...]}
    jp_helper = JapaneseHelper()
    barcode_pattern = re.compile(r'\d{8}-\d{3}')

    for idx, path in enumerate(meisai_paths):
        log_fn(f"  📋 Reading barcode map from: {os.path.basename(path)}\n")
        try:
            # อัพเดท progress ระหว่าง analyze 注文明細票
            if progress_callback:
                p = base_progress + (progress_range * idx / len(meisai_paths))
                progress_callback(p)
            result = analyze_invoice(path)
            if progress_callback:
                p = base_progress + (progress_range * (idx + 1) / len(meisai_paths))
                progress_callback(p)
            tables = getattr(result, "tables", None) or []
            _, best_table = pick_best_items_table(result, jp_helper)
            if not best_table:
                log_fn(f"  ⚠ No table found in {os.path.basename(path)}\n")
                continue

            n_rows = best_table.row_count
            n_cols = best_table.column_count

            # Build flat grid
            grid = [['' for _ in range(n_cols)] for _ in range(n_rows)]
            for tcell in best_table.cells:
                r, c = tcell.row_index, tcell.column_index
                rs = max(getattr(tcell, "row_span", 1) or 1, 1)
                cs = max(getattr(tcell, "column_span", 1) or 1, 1)
                val = jp_helper.normalize_text(tcell.content or "")
                for dr in range(rs):
                    for dc in range(cs):
                        rr, cc = r + dr, c + dc
                        if rr < n_rows and cc < n_cols:
                            grid[rr][cc] = val if (dr == 0 and dc == 0) else ""

            # Find header row
            hdr_row = 0
            for r in range(min(5, n_rows)):
                row_text = " ".join(grid[r])
                if any(k in row_text for k in ["品名", "納期", "数量"]):
                    hdr_row = r
                    break

            # Find barcode column (column ที่ค่ามี pattern XXXXXXXX-XXX)
            barcode_col = -1
            for c in range(n_cols - 1, -1, -1):
                for r in range(hdr_row + 1, min(hdr_row + 5, n_rows)):
                    if barcode_pattern.search(grid[r][c]):
                        barcode_col = c
                        break
                if barcode_col >= 0:
                    break

            if barcode_col < 0:
                log_fn(f"  ⚠ Barcode column not found\n")
                continue

            # Find 数量 column จาก header
            qty_col = -1
            hdr_text = [grid[hdr_row][c] for c in range(n_cols)]
            for c, h in enumerate(hdr_text):
                if any(k in h for k in ["数量", "員数", "個数"]):
                    qty_col = c
                    break

            # Build map: (品番 + 数量) → barcode
            # key = "品番__数量"  เช่น "60008-72240-0__5"
            # ถ้า key ซ้ำ → mark เป็น duplicate ให้คนจัดการเอง
            part_col = 0
            count = 0
            duplicate_keys = set()

            for r in range(hdr_row + 1, n_rows):
                part_raw = grid[r][part_col].strip()
                barcode_raw = grid[r][barcode_col].strip()
                qty_raw = grid[r][qty_col].strip() if qty_col >= 0 else ""

                if not part_raw or not barcode_raw:
                    continue
                if not barcode_pattern.search(barcode_raw):
                    continue

                part_code = extract_part_no(part_raw)
                if not part_code:
                    continue

                # normalize 数量 — ดึงแค่ตัวเลขนำหน้า เช่น "5¥250-" → "5", "5 個" → "5"
                qty_stripped = re.sub(r'[,\s]', '', qty_raw) if qty_raw else ""
                qty_num = re.search(r'^\d+', qty_stripped)
                qty_val = qty_num.group() if qty_num else ""
                composite_key = f"{part_code}__{qty_val}"
                barcode_code = barcode_pattern.search(barcode_raw).group()

                if composite_key in barcode_map:
                    duplicate_keys.add(composite_key)  # mark ว่าซ้ำ
                    barcode_map[composite_key].append(barcode_code)
                else:
                    barcode_map[composite_key] = [barcode_code]
                count += 1

            unique = count - sum(len(v) - 1 for k, v in barcode_map.items() if k in duplicate_keys)
            log_fn(f"  ✓ {count} mappings — unique: {unique} / ⚠ duplicate key: {len(duplicate_keys)}\n")
            if duplicate_keys:
                for dk in list(duplicate_keys)[:5]:
                    log_fn(f"    ⚠ duplicate: {dk}\n")

        except Exception as e:
            log_fn(f"  ✗ Error reading barcode map: {e}\n")

    return barcode_map



# ======================
# Smart Cell Value Converter
# ======================
def smart_cell_value(text, jp_helper=None):
    """
    แปลง string → Python type ที่เหมาะสมก่อนเขียน Excel
    - วันที่ญี่ปุ่น → datetime object  (Excel filter/sort ได้)
    - ตัวเลข ¥240, 3,800, 240- → int/float  (SUM ได้)
    - Part No. และ string อื่นๆ → string เดิม
    """
    from datetime import datetime as _dt
    if not isinstance(text, str):
        return text
    s = text.strip()
    if not s:
        return s

    # 1. ลองแปลงวันที่
    if jp_helper:
        parsed = jp_helper.parse_date(s)
        if re.match(r'^\d{4}-\d{2}-\d{2}$', parsed):
            try:
                return _dt.strptime(parsed, "%Y-%m-%d")
            except Exception:
                pass

    # 2. แปลงตัวเลขแบบญี่ปุ่น (strip ¥ ￥ , space)
    num_s = re.sub(r'[¥￥,\s]', '', s)
    # trailing - แบบญี่ปุ่น เช่น "240-" → ตัดออก (ไม่ใช่ negative)
    if re.match(r'^\d[\d.]*-$', num_s):
        num_s = num_s[:-1]
    # แปลงเป็น int เฉพาะเมื่อไม่มี 0 นำหน้า (ป้องกัน 04743903 กลายเป็น 4743903)
    if re.match(r'^[1-9]\d*$', num_s):
        return int(num_s)
    if re.match(r'^[1-9]\d*\.\d+$|^0\.\d+$', num_s):
        return float(num_s)

    return s

def export_many_to_one_excel(pdf_paths: list, out_xlsx: str, log_fn, progress_callback=None, barcode_mode: bool = False, meisai_override=None) -> tuple:
    """
    Combine many PDFs into ONE Excel
    Sheet 1: Summary  - 1 row/file (FileName + key header fields จาก JSON)
    Sheet 2: All_Items - รวม RAW tables จากทุกไฟล์ (เหมือน Separate Mode)
    """
    log_fn(f"\n{'='*60}\n")
    log_fn(f"Combining {len(pdf_paths)} PDFs to ONE Excel\n")
    log_fn(f"{'='*60}\n\n")

    wb = Workbook()
    wb.remove(wb.active)
    jp_helper = JapaneseHelper()

    # ── Build barcode map ถ้า barcode_mode เปิด ──────────────────
    barcode_map = {}
    meisai_paths = []
    delivery_paths = []
    if barcode_mode:
        print(f"DEBUG barcode_mode=True, pdf_paths={len(pdf_paths)}, meisai_override={len(meisai_override) if meisai_override else None}")
        for p in pdf_paths:
            is_meisai = (p in meisai_override) if meisai_override is not None else detect_meisai_file(p)
            print(f"DEBUG  {'[M]' if is_meisai else '[D]'} {os.path.basename(p)}")
            # ใช้ meisai_override จาก UI ถ้ามี ไม่งั้น fallback detect จากชื่อไฟล์
            is_meisai = (p in meisai_override) if meisai_override is not None else detect_meisai_file(p)
            if is_meisai:
                meisai_paths.append(p)
            else:
                delivery_paths.append(p)
        print(f"DEBUG Barcode mode: {len(meisai_paths)} meisai, {len(delivery_paths)} delivery")
        log_fn(f"Barcode mode: {len(meisai_paths)} 注文明細票, {len(delivery_paths)} 納品書\n\n")
        try:
            if meisai_paths:
                n_total = len(pdf_paths) + len(meisai_paths)
                meisai_range = len(meisai_paths) / n_total
                barcode_map = build_barcode_map(
                    meisai_paths, log_fn,
                    progress_callback=progress_callback,
                    base_progress=0.0,
                    progress_range=meisai_range
                )
                dup_count = sum(1 for v in barcode_map.values() if len(v) > 1)
                log_fn(f"✓ Barcode map: {len(barcode_map)} unique Part No. ({dup_count} duplicates)\n\n")
        except Exception as _bce:
            print(f"DEBUG build_barcode_map ERROR: {_bce}")
            import traceback; traceback.print_exc()
        # ถ้า barcode_mode แต่ไม่มี meisai ตรวจไปแล้วตั้งแต่ UI
        # process เฉพาะ delivery files
        pdf_paths = delivery_paths if delivery_paths else pdf_paths

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="366092")
    hdr_font = Font(name="MS Gothic", size=11, bold=True, color="FFFFFF")
    base_font = Font(name="MS Gothic", size=10)

    # ── Summary sheet ──────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    SUMMARY_FIELDS = ["FileName", "ScanDate", "InvoiceDate", "Page", "InvoiceId", "VendorName", "CustomerName", "InvoiceTotal"]
    for c, name in enumerate(SUMMARY_FIELDS, 1):
        label = JP_LABELS.get(name, name)
        cell = ws_sum.cell(1, c, label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    if barcode_mode:
        sum_bc_col = len(SUMMARY_FIELDS) + 1
        for offset, (label, color) in enumerate([("照合✅", "1B5E20"), ("未一致⚠", "B71C1C")]):
            cell = ws_sum.cell(1, sum_bc_col + offset, label)
            cell.font = Font(name="MS Gothic", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
    ws_sum.row_dimensions[1].height = 22

    # ── All_Items sheet ─────────────────────────────────────────
    ws_all = wb.create_sheet("All_Items")
    items_header_written = False   # เขียน header row ครั้งแรกจากไฟล์แรก
    items_cur_row = 2              # row ที่จะเขียน data ถัดไป (row 1 = header)
    line_no = 1                    # Line counter ต่อเนื่องทุกไฟล์

    success = 0
    failed = 0
    sum_row = 2                    # row ใน Summary sheet

    meisai_count = len(meisai_paths) if barcode_mode else 0
    n_total = len(pdf_paths) + meisai_count
    base_offset = meisai_count / n_total if n_total else 0
    remaining_range = 1.0 - base_offset

    for idx, pdf_path in enumerate(pdf_paths, 1):
        try:
            fname = os.path.basename(pdf_path)
            log_fn(f"[{idx}/{len(pdf_paths)}] {fname}\n")
            file_matched = 0
            file_unmatched = 0

            result = analyze_invoice(pdf_path)

            # อัพเดท progress หลัง analyze เสร็จแต่ละไฟล์
            if progress_callback:
                p = base_offset + remaining_range * (idx / len(pdf_paths))
                # หักไว้ 5% สำหรับขั้นตอน save Excel ท้ายสุด
                progress_callback(min(p, 0.95))
            header, _ = extract_fields(result, jp_helper)

            # ── ดึง Page จาก Table 1 (2-column key/value table) ──
            # Table 1 มีโครงสร้างแบบ:
            # row 0: "Page" | "31/32"
            # row 1: "No."  | "28966"
            page_value = ""
            tables = getattr(result, "tables", None) or []
            for table in tables:
                if table.column_count < 2:
                    continue
                # Build rows
                rows = {}
                for tcell in table.cells:
                    rows.setdefault(tcell.row_index, {})[tcell.column_index] = (
                        jp_helper.normalize_text(tcell.content or "")
                    )
                all_text = " ".join(v for row in rows.values() for v in row.values()).lower()
                # เป็น metadata table ถ้า: row น้อย และมี "page"
                if table.row_count <= 5 and "page" in all_text:
                    for row_data in rows.values():
                        # col 0 = key, col 1 = value
                        key = row_data.get(0, "").lower()
                        val = row_data.get(1, "")
                        if "page" in key and val:
                            page_value = val
                            break
                    break

            # ── 1. เขียน Summary row ──
            # ดึงวันที่จาก FileName เช่น 20260220093545897_0001.pdf → 2026-02-20
            from datetime import datetime as _dt
            scan_date = None
            _fm = re.match(r'(\d{4})(\d{2})(\d{2})\d+', fname)
            if _fm:
                try:
                    scan_date = _dt(int(_fm.group(1)), int(_fm.group(2)), int(_fm.group(3)))
                except Exception:
                    scan_date = None

            row_fill = PatternFill("solid", fgColor="F2F2F2" if sum_row % 2 == 0 else "FFFFFF")
            this_sum_row = sum_row  # จำไว้เพื่อกลับมาเติม barcode count หลัง data loop เสร็จ
            for c, field in enumerate(SUMMARY_FIELDS, 1):
                if field == "FileName":
                    val = fname
                elif field == "ScanDate":
                    val = scan_date  # datetime object → Excel date
                elif field == "Page":
                    val = page_value
                else:
                    raw = str(header.get(field, "") or "")
                    # InvoiceId = string เสมอ ไม่แปลงเป็นตัวเลข + ตัด "No" นำหน้า
                    if field == "InvoiceId":
                        val = re.sub(r'^[Nn][Oo]\.?\s*', '', raw).strip()
                    else:
                        val = smart_cell_value(raw, jp_helper) if raw else ""
                cell = ws_sum.cell(sum_row, c, val)
                cell.font = base_font
                cell.border = border
                cell.fill = row_fill
                if isinstance(val, _dt):
                    cell.number_format = "YYYY/MM/DD"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif isinstance(val, (int, float)):
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
            ws_sum.row_dimensions[sum_row].height = 18
            sum_row += 1
            # หมายเหตุ: barcode count (file_matched / file_unmatched) จะถูกเติมหลัง data loop
            # ด้านล่าง โดยอ้างอิง this_sum_row

            # ── 2. ดึง best RAW table (เหมือน Separate Mode) ──
            _, best_table = pick_best_items_table(result, jp_helper)
            if not best_table:
                log_fn(f"  ⚠ No table found\n")
                success += 1
                continue

            # Build FLAT GRID (expand row_span/col_span) เหมือน create_layout_sheet
            n_rows_t = best_table.row_count
            n_cols = best_table.column_count
            grid = [["" for _ in range(n_cols)] for _ in range(n_rows_t)]
            for tcell in best_table.cells:
                r, c = tcell.row_index, tcell.column_index
                rs = max(getattr(tcell, "row_span", 1) or 1, 1)
                cs = max(getattr(tcell, "column_span", 1) or 1, 1)
                val = jp_helper.normalize_text(tcell.content or "")
                for dr in range(rs):
                    for dc in range(cs):
                        rr, cc = r + dr, c + dc
                        if rr < n_rows_t and cc < n_cols:
                            grid[rr][cc] = val if (dr == 0 and dc == 0) else grid[rr][cc]

            # Convert grid back to rows dict format
            rows = {r: {c: grid[r][c] for c in range(n_cols)} for r in range(n_rows_t)}
            sorted_rows = list(range(n_rows_t))

            # หา header row
            hdr_row_idx = 0
            for r in range(min(5, n_rows_t)):
                row_text = " ".join(grid[r]).lower()
                if any(k in row_text for k in ["品名", "数量", "単価", "金額", "出荷", "伝票"]):
                    hdr_row_idx = r
                    break

            col_headers = {c: grid[hdr_row_idx][c] for c in range(n_cols)}

            # ── เขียน header row ครั้งแรก (จากไฟล์แรกที่มีข้อมูล) ──
            if not items_header_written:
                # Col 1=Line, Col 2=FileName, Col 3=伝票番号, Col 4=ScanDate, Col 5+=data columns
                for c_idx, label in enumerate(["Line", "FileName", "伝票番号", "ScanDate"], 1):
                    cell = ws_all.cell(1, c_idx, label)
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border

                for c_idx in range(n_cols):
                    col_name = col_headers.get(c_idx, f"Col{c_idx}")
                    cell = ws_all.cell(1, c_idx + 5, col_name)  # +5
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border

                # เพิ่ม Barcode + Match Status columns ท้ายสุด
                if barcode_mode:
                    bc_col = n_cols + 5
                    for col_offset, (label, color) in enumerate([
                        ("照合Part No.", "5C4033"),
                        ("Barcode",      "7B3F00"),
                        ("照合結果",     "1B5E20"),
                    ]):
                        cell = ws_all.cell(1, bc_col + col_offset, label)
                        cell.font = Font(name="MS Gothic", size=10, bold=True, color="FFFFFF")
                        cell.fill = PatternFill("solid", fgColor=color)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border

                ws_all.row_dimensions[1].height = 22
                items_header_written = True
                log_fn(f"  → Header columns: {n_cols} columns\n")

            # ── เขียน data rows ──
            data_rows = [r for r in sorted_rows if r != hdr_row_idx]
            written = 0
            for r in data_rows:
                row_data = rows.get(r, {})

                # Skip rows ที่ว่างหรือเป็น subtotal
                row_text = " ".join(str(v) for v in row_data.values())
                if not row_text.strip():
                    continue
                if any(k in row_text for k in ["小計", "合計", "総計", "頁計"]):
                    continue

                row_fill = PatternFill("solid", fgColor="F2F2F2" if items_cur_row % 2 == 0 else "FFFFFF")

                # Col 1 = Line number
                cell = ws_all.cell(items_cur_row, 1, line_no)
                cell.font = Font(name="MS Gothic", size=10, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="top")
                cell.border = border
                cell.fill = row_fill

                # Col 2 = FileName (แสดงเฉพาะ row แรกของแต่ละไฟล์)
                cell = ws_all.cell(items_cur_row, 2, fname if written == 0 else "")
                cell.font = Font(name="MS Gothic", size=10, color="595959")
                cell.alignment = Alignment(horizontal="left", vertical="top")
                cell.border = border
                cell.fill = row_fill

                # Col 3 = 伝票番号 (ทุก row — เพื่อจับชุดเอกสาร)
                invoice_id = re.sub(r'^[Nn][Oo]\.?\s*', '', str(header.get("InvoiceId", "") or "")).strip()
                cell = ws_all.cell(items_cur_row, 3, invoice_id)
                cell.font = base_font
                cell.border = border
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal="center", vertical="top")

                # Col 4 = ScanDate (ทุก row)
                from datetime import datetime as _dt
                cell = ws_all.cell(items_cur_row, 4, scan_date)
                cell.font = base_font
                cell.border = border
                cell.fill = row_fill
                if scan_date:
                    cell.number_format = "YYYY/MM/DD"
                    cell.alignment = Alignment(horizontal="center", vertical="top")

                # Col 5+ = data — ใช้ค่าดิบตรงๆ ไม่ผ่าน smart_cell_value
                # เพราะค่าเช่น '0600 2026/4/2' จะถูก smart_cell_value ลบ space ทิ้ง
                for c_idx in range(n_cols):
                    raw_val = row_data.get(c_idx, "")
                    val = str(raw_val) if raw_val is not None else ""
                    cell = ws_all.cell(items_cur_row, c_idx + 5, val)  # +5
                    cell.font = base_font
                    cell.border = border
                    cell.fill = row_fill

                    col_name = str(col_headers.get(c_idx, ""))
                    if any(k in col_name for k in ["単価", "金額", "重量", "員数", "数量"]):
                        cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
                    elif any(k in col_name for k in ["出荷日", "納期", "日付"]):
                        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

                # เพิ่ม Barcode + Match Status
                if barcode_mode:
                    bc_col = n_cols + 5
                    part_raw = str(row_data.get(0, "")).strip()
                    part_code = extract_part_no(part_raw)

                    # หา 数量 จาก row นี้ — ดูจาก col ที่ header มีคำว่า 数量
                    qty_val = ""
                    for c_i in range(n_cols):
                        h = str(col_headers.get(c_i, ""))
                        if any(k in h for k in ["数量", "員数", "個数", "数", "量"]):
                            qty_raw2 = re.sub(r'[,\s]', '', str(row_data.get(c_i, "")).strip())
                            qty_num2 = re.search(r'^\d+', qty_raw2)
                            qty_val = qty_num2.group() if qty_num2 else qty_raw2
                            break

                    composite_key = f"{part_code}__{qty_val}"
                    barcode_list = barcode_map.get(composite_key, [])

                    # ตรวจว่า key นี้ duplicate ไหม (ก่อน pop)
                    is_duplicate = len(barcode_list) > 1

                    if barcode_list:
                        barcode_val = barcode_list.pop(0)  # pop เสมอ ไม่ว่าจะ duplicate หรือเปล่า
                        if is_duplicate:
                            match_status = "✅ 自動照合 (重複)"
                            status_color = "C8E6C9"   # green เหมือนกัน เพราะจับคู่ได้
                        else:
                            match_status = "✅ 自動照合"
                            status_color = "C8E6C9"   # green
                    else:
                        barcode_val = ""
                        match_status = "❌ 未一致"
                        status_color = "FFCDD2"   # red

                    # col+0: Part No. ที่ใช้จับคู่
                    c0 = ws_all.cell(items_cur_row, bc_col, part_code)
                    c0.font = base_font
                    c0.border = border
                    c0.fill = PatternFill("solid", fgColor=status_color)
                    c0.alignment = Alignment(horizontal="left", vertical="top")

                    # col+1: Barcode value
                    c1 = ws_all.cell(items_cur_row, bc_col + 1, barcode_val)
                    c1.font = base_font
                    c1.border = border
                    c1.fill = PatternFill("solid", fgColor=status_color)
                    c1.alignment = Alignment(horizontal="left", vertical="top")

                    # col+2: Match Status
                    c2 = ws_all.cell(items_cur_row, bc_col + 2, match_status)
                    c2.font = Font(name="MS Gothic", size=10, bold=True)
                    c2.border = border
                    c2.fill = PatternFill("solid", fgColor=status_color)
                    c2.alignment = Alignment(horizontal="center", vertical="top")

                items_cur_row += 1
                line_no += 1
                written += 1
                if barcode_mode:
                    if match_status.startswith("✅"):
                        file_matched += 1
                    else:
                        file_unmatched += 1

            if barcode_mode:
                # เติม barcode count ใน Summary row หลัง data loop เสร็จ (ตัวเลขถูกต้องแล้ว)
                ws_sum.cell(this_sum_row, sum_bc_col, file_matched)
                ws_sum.cell(this_sum_row, sum_bc_col).fill = PatternFill("solid", fgColor="C8E6C9")
                ws_sum.cell(this_sum_row, sum_bc_col).border = border
                ws_sum.cell(this_sum_row, sum_bc_col).alignment = Alignment(horizontal="center", vertical="center")
                unmatched_color = "FFCDD2" if file_unmatched > 0 else "F5F5F5"
                ws_sum.cell(this_sum_row, sum_bc_col + 1, file_unmatched)
                ws_sum.cell(this_sum_row, sum_bc_col + 1).fill = PatternFill("solid", fgColor=unmatched_color)
                ws_sum.cell(this_sum_row, sum_bc_col + 1).border = border
                ws_sum.cell(this_sum_row, sum_bc_col + 1).alignment = Alignment(horizontal="center", vertical="center")
                log_fn(f"  ✓ {written} 件追加 — 照合: ✅{file_matched} 件 / ❌{file_unmatched} 件未一致\n")
            else:
                log_fn(f"  ✓ Added {written} items\n")
            success += 1

        except Exception as e:
            failed += 1
            log_fn(f"  ✗ Error: {e}\n")

    # ── Format Summary sheet ──────────────────────────────────
    for c, name in enumerate(SUMMARY_FIELDS, 1):
        max_len = len(JP_LABELS.get(name, name))
        for r in range(2, sum_row):
            v = ws_sum.cell(r, c).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws_sum.column_dimensions[get_column_letter(c)].width = min(max_len + 3, 50)
    ws_sum.freeze_panes = "A2"

    # ── Sort + Format All_Items sheet ───────────────────────
    if items_header_written:
        # Sort ตาม 伝票番号 (col 3) ก่อน format
        # อ่านทุก data row ออกมา sort แล้วเขียนกลับ
        data_rows_all = []
        max_col = ws_all.max_column
        for r in range(2, items_cur_row):
            row_vals = []
            row_styles = []
            for c in range(1, max_col + 1):
                cell = ws_all.cell(r, c)
                row_vals.append(cell.value)
                row_styles.append({
                    'font': copy(cell.font),
                    'border': copy(cell.border),
                    'fill': copy(cell.fill),
                    'alignment': copy(cell.alignment),
                    'number_format': cell.number_format,
                })
            data_rows_all.append((row_vals, row_styles))

        # sort key = col 3 (伝票番号), None ไปท้าย
        def sort_key(item):
            v = item[0][2]  # col index 2 = col 3
            if v is None or v == "":
                return (1, "")
            return (0, str(v))

        data_rows_all.sort(key=sort_key)

        # เขียนกลับ พร้อม re-zebra stripe
        for i, (row_vals, row_styles) in enumerate(data_rows_all):
            excel_row = i + 2
            zebra = PatternFill("solid", fgColor="F2F2F2" if excel_row % 2 == 0 else "FFFFFF")
            for c_idx, (val, style) in enumerate(zip(row_vals, row_styles)):
                cell = ws_all.cell(excel_row, c_idx + 1, val)
                cell.font = style['font']
                cell.border = style['border']
                cell.fill = zebra
                cell.alignment = style['alignment']
                cell.number_format = style['number_format']

        log_fn(f"✓ All_Items sorted by 伝票番号\n")

        # Auto column width
        for c in range(1, max_col + 1):
            max_len = 0
            for r in range(1, items_cur_row):
                v = ws_all.cell(r, c).value
                if v:
                    max_len = max(max_len, len(str(v)))
            col_name = str(ws_all.cell(1, c).value or "")
            width = min(max(8, max_len * 1.2), 55)
            if any(k in col_name for k in ["品名", "摘要", "部品名"]):
                width = max(width, 35)
            ws_all.column_dimensions[get_column_letter(c)].width = width
        ws_all.freeze_panes = "A2"

    wb.save(out_xlsx)
    log_fn(f"\n{'='*60}\n")
    log_fn(f"✓ Saved: {os.path.basename(out_xlsx)}\n")
    log_fn(f"  Total rows: {line_no - 1}\n")
    log_fn(f"  Success: {success}, Failed: {failed}\n")
    log_fn(f"{'='*60}\n\n")
    return success, failed




def _split_part_name(raw: str):
    """แยก col0 '60008-72240-0 リンク、ベースプレート' → ('60008-72240-0', 'リンク、ベースプレート')"""
    raw = (raw or "").strip()
    parts = raw.split(None, 1)
    if not parts:
        return "", ""
    return parts[0], (parts[1] if len(parts) > 1 else "")


# ──────────────────────────────────────────────────────────────
# NEW Barcode-mode export: meisai (layout+barcodes) ↔ delivery
# ──────────────────────────────────────────────────────────────
def export_barcode_mode(meisai_paths, delivery_paths, out_xlsx, log_fn, progress_callback=None):
    """
    โหมด barcode ใหม่ — ใช้ barcode_match:
      1. meisai: prebuilt-layout + BARCODES → จับ barcode เข้าบรรทัดด้วยพิกัด Y
      2. จับคู่เอกสาร meisai↔delivery ด้วยชุด Part No.
      3. แมตช์บรรทัดด้วย Part No. (+ qty/price ตัดตัวซ้ำ)
    Output 4 ชีต: 納品一覧(Summary) / 明細 / 要確認 / 照合残り(Reconcile)
    """
    jp = JapaneseHelper()
    norm = jp.normalize_text
    pick = lambda r: pick_best_items_table(r, jp)

    # ── Light clean palette ──
    C_HEADER = "2563EB"      # น้ำเงินสด
    C_HEADER_TX = "FFFFFF"
    C_GRID = "E5E7EB"        # เส้นตารางเทาอ่อน
    C_ZEBRA = "F8FAFC"
    C_OK = "DCFCE7"          # เขียวอ่อน
    C_REVIEW = "FEF3C7"      # เหลืองอ่อน
    C_NOBC = "F1F5F9"        # เทา
    C_UNMATCH = "FEE2E2"     # แดงอ่อน
    thin = Side(style="thin", color=C_GRID)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font = Font(name="MS Gothic", size=10, bold=True, color=C_HEADER_TX)
    base_font = Font(name="MS Gothic", size=10, color="1F2937")
    hdr_fill = PatternFill("solid", fgColor=C_HEADER)

    STATUS_DISP = {
        bm.STATUS_OK:        ("✅ 照合",       C_OK),
        bm.STATUS_REVIEW:    ("⚠ 要確認",     C_REVIEW),
        bm.STATUS_NO_BARCODE:("― barcodeなし", C_NOBC),
        bm.STATUS_UNMATCHED: ("❌ 未一致",     C_UNMATCH),
    }

    def style_header(ws, labels, row=1, start=1):
        for i, lab in enumerate(labels, start):
            c = ws.cell(row, i, lab)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        ws.row_dimensions[row].height = 24

    wb = Workbook()
    wb.remove(wb.active)

    n_total = max(1, len(meisai_paths) + len(delivery_paths))
    done = [0]
    def tick():
        done[0] += 1
        if progress_callback:
            progress_callback(min(0.95, done[0] / n_total))

    def scan_date_of(path):
        m = re.match(r"(\d{4})(\d{2})(\d{2})\d+", os.path.basename(path))
        if m:
            try:
                return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except Exception:
                return None
        return None

    meisai_fail = 0
    read_fail = 0

    # ── Phase 1: index meisai ─────────────────────────────────
    log_fn(f"\n{'='*60}\n📋 STEP 1 — 注文明細票 {len(meisai_paths)} ไฟล์ (อ่าน barcode)\n{'='*60}\n")
    meisai_indexes = {}
    for mp in meisai_paths:
        log_fn(f"  • {os.path.basename(mp)}\n")
        try:
            res = bm.analyze_layout_with_barcodes(ENDPOINT, KEY, mp, log_fn=log_fn)
            pz = bm.decode_barcodes_pyzbar(mp)          # อ่านจากรูปจริงเสริม Azure
            idx = bm.index_meisai(res, norm, pick, pyzbar_pages=pz)
            meisai_indexes[mp] = idx
            log_fn(f"      barcode {idx['bc_matched']}/{idx['bc_total']} จับเข้าบรรทัดได้\n")
            if idx["bc_total"] == 0:
                log_fn(f"      ⚠ ไฟล์นี้ไม่พบ barcode เลย — แน่ใจว่าเป็น 注文明細票?\n")
        except Exception as e:
            meisai_fail += 1
            log_fn(f"      ✗✗ 読取失敗 (meisai): {type(e).__name__}: {e}\n")
        tick()

    # ── Phase 2: read deliveries ──────────────────────────────
    log_fn(f"\n{'='*60}\n📄 STEP 2 — 納品書 {len(delivery_paths)} ไฟล์ (อ่านตาราง)\n{'='*60}\n")
    delivery_tables = {}
    delivery_nos = {}        # dp -> เลขใบ (No.) สำหรับรวมหน้า
    for dp in delivery_paths:
        log_fn(f"  • {os.path.basename(dp)}\n")
        try:
            res = analyze_invoice(dp, log_fn=log_fn)
            dtab = bm.read_items_table(res, norm, pick)
            dtab["_failed"] = False
            delivery_tables[dp] = dtab
            # ดึงเลขใบ (No.) จาก header — ใช้ผลที่อ่านอยู่แล้ว ไม่กิน Azure เพิ่ม
            try:
                _hdr, _ = extract_fields(res, jp)
                no = re.sub(r"^[Nn][Oo]\.?\s*", "", str(_hdr.get("InvoiceId", "") or "")).strip()
            except Exception:
                no = ""
            delivery_nos[dp] = no
            log_fn(f"      {len(dtab['rows'])} บรรทัด · No.{no or '—'}\n")
            if not dtab["rows"]:
                log_fn(f"      ⚠ ตารางว่าง — อ่านตารางไม่เจอ\n")
        except Exception as e:
            read_fail += 1
            log_fn(f"      ✗✗ 読取失敗 (delivery): {type(e).__name__}: {e}\n")
            delivery_tables[dp] = {"headers": [], "rows": [], "roles": {}, "_failed": True}
            delivery_nos[dp] = ""
        tick()

    # ── Phase 2.5: รวมไฟล์ที่เลขใบ (No.) เดียวกัน = ใบส่งของหลายหน้า ─────
    delivery_groups = []   # [{key,no,files[],rows[],row_files[],roles{},partset,failed}]
    _by_no = {}
    for dp in delivery_paths:
        dtab = delivery_tables[dp]
        no = delivery_nos.get(dp, "")
        roles = dtab.get("roles", {})
        key = no if no else f"__file__::{dp}"   # ไม่มีเลข = ใบเดี่ยว
        # ถ้าคอลัมน์ (roles) ของไฟล์ใหม่ไม่ตรงกับกลุ่มเดิม → แยกเป็นใบเดี่ยว กันข้อมูลเพี้ยน
        g = _by_no.get(key)
        if g is not None and g["roles"] and roles and roles != g["roles"]:
            key = f"__file__::{dp}"
            g = None
        if g is None:
            g = {"key": key, "no": no, "files": [], "rows": [], "row_files": [],
                 "roles": roles or {}, "partset": set(), "failed": False}
            _by_no[key] = g
            delivery_groups.append(g)
        if not g["roles"] and roles:
            g["roles"] = roles
        g["files"].append(dp)
        base = os.path.basename(dp)
        rr = dtab.get("roles", {})
        for row in dtab.get("rows", []):
            g["rows"].append(row)
            g["row_files"].append(base)
            p = bm.extract_part_no(row[rr["part"]]) if rr.get("part", -1) >= 0 and rr["part"] < len(row) else ""
            if p:
                g["partset"].add(p)
        if dtab.get("_failed"):
            g["failed"] = True

    # ── Phase 3: pair documents (จับคู่ที่ระดับ "ใบ" หลังรวมหน้าแล้ว) ──
    group_partsets = {g["key"]: g["partset"] for g in delivery_groups}
    pairing = bm.pair_documents(meisai_indexes, group_partsets)
    log_fn(f"\n{'='*60}\n🔗 STEP 3 — จับคู่เอกสาร ({len(delivery_groups)} ใบ จาก {len(delivery_paths)} ไฟล์)\n{'='*60}\n")
    for g in delivery_groups:
        pr = pairing.get(g["key"], {})
        mp = pr.get("meisai")
        tag = f"No.{g['no']}" if g["no"] else os.path.basename(g["files"][0])
        if len(g["files"]) > 1:
            tag += f" [{len(g['files'])}頁]"
        if mp:
            log_fn(f"  {tag} ↔ {os.path.basename(mp)} (一致度 {pr['confidence']:.0%}, {pr['overlap']} parts)\n")
        else:
            log_fn(f"  {tag} ↔ ❌ ไม่พบคู่\n")

    # ── Sheets ────────────────────────────────────────────────
    ws_sum = wb.create_sheet("納品一覧")
    SUM_COLS = ["納品書ファイル", "スキャン日", "対応 明細票", "一致度",
                "✅照合", "⚠要確認", "barcodeなし", "❌未一致", "行数", "判定"]
    style_header(ws_sum, SUM_COLS)

    ws_det = wb.create_sheet("明細")
    DET_COLS = ["No", "納品書", "納期", "品番", "品名", "数量", "単位", "単価", "金額",
                "備考", "Barcode", "照合状態", "メモ"]
    style_header(ws_det, DET_COLS)
    ws_det.freeze_panes = "A2"

    ws_rev = wb.create_sheet("要確認")
    style_header(ws_rev, ["No", "納品書", "品番", "品名", "数量", "単価", "状態", "候補barcode", "メモ"])
    ws_rev.freeze_panes = "A2"

    # ── Phase 4: match + write ────────────────────────────────
    det_row = 2
    rev_row = 2
    line_no = 1
    sum_row = 2
    success = 0
    failed = 0

    for grp in delivery_groups:
        try:
            files = grp["files"]
            rows = grp["rows"]
            roles = grp["roles"]
            row_files = grp["row_files"]
            sdate = scan_date_of(files[0])
            if len(files) > 1:
                grp_disp = f"{os.path.basename(files[0])} +{len(files)-1}頁 (No.{grp['no']})"
            else:
                grp_disp = os.path.basename(files[0])
            pr = pairing.get(grp["key"], {})
            mp = pr.get("meisai")
            midx = meisai_indexes.get(mp) if mp else None

            cnt = {bm.STATUS_OK: 0, bm.STATUS_REVIEW: 0, bm.STATUS_NO_BARCODE: 0, bm.STATUS_UNMATCHED: 0}

            match_results = bm.match_delivery(rows, roles, midx)

            for ri, row in enumerate(rows):
                fname = row_files[ri]

                def g(role):
                    c = roles.get(role, -1)
                    return row[c] if (c is not None and 0 <= c < len(row)) else ""

                part_raw = g("part")
                code_disp, name_disp = _split_part_name(part_raw)
                qty = g("qty"); unit = g("unit"); price = g("price")
                amount = g("amount"); remark = g("remark")

                res = match_results[ri]
                status = res["status"]
                cnt[status] = cnt.get(status, 0) + 1
                disp, color = STATUS_DISP[status]
                fill = PatternFill("solid", fgColor=color)

                # 納期 จากแถว meisai ที่จับคู่ได้ (แปลงวันที่ญี่ปุ่น → date)
                nouki_raw = res.get("nouki", "")
                nouki_val = ""
                if nouki_raw:
                    ds = jp.parse_date(nouki_raw)
                    if re.match(r"^\d{4}-\d{2}-\d{2}$", ds):
                        try:
                            nouki_val = datetime.strptime(ds, "%Y-%m-%d")
                        except Exception:
                            nouki_val = nouki_raw
                    else:
                        nouki_val = nouki_raw

                vals = [line_no, fname if line_no else "", nouki_val, code_disp, name_disp,
                        qty, unit, price, amount, remark, res["barcode"], disp,
                        res.get("note", "")]
                for ci, v in enumerate(vals, 1):
                    c = ws_det.cell(det_row, ci, v)
                    c.font = base_font
                    c.border = border
                    c.fill = fill
                    if ci in (6, 8, 9):  # 数量/単価/金額 ชิดขวา
                        c.alignment = Alignment(horizontal="right", vertical="center")
                    elif ci in (1, 3, 7, 12):
                        c.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    if ci == 3 and isinstance(nouki_val, datetime):
                        c.number_format = "YYYY/MM/DD"
                det_row += 1
                line_no += 1

                # 要確認 sheet เฉพาะที่ต้องตรวจ
                if status in (bm.STATUS_REVIEW, bm.STATUS_UNMATCHED):
                    rvals = [line_no - 1, fname, code_disp, name_disp, qty, price, disp,
                             ", ".join(res.get("candidates", [])), res.get("note", "")]
                    for ci, v in enumerate(rvals, 1):
                        c = ws_rev.cell(rev_row, ci, v)
                        c.font = base_font
                        c.border = border
                        c.fill = PatternFill("solid", fgColor=color)
                        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    rev_row += 1

            # Summary row (ต่อ "ใบ" = หลังรวมหน้าแล้ว)
            if grp.get("failed"):
                pair_disp = "⚠ 読取失敗"
            elif mp:
                pair_disp = os.path.basename(mp)
            else:
                pair_disp = "❌ なし"

            conf = pr.get("confidence", 0.0) if mp else 0.0
            n_unmatch = cnt[bm.STATUS_UNMATCHED]
            n_review = cnt[bm.STATUS_REVIEW]

            # ── 判定 (cross-check แบบไม่กระทบ match — แค่เตือนใบที่น่าสงสัย) ──
            # ใช้ "coverage" (一致度) + จำนวน ❌ ไม่ใช่ "ยอดรวมต้องเท่ากัน"
            # (meisai = ทั้งออเดอร์, delivery = ส่งบางส่วน → ยอดไม่มีทางเท่า)
            if grp.get("failed"):
                verdict, vcolor, conf_color = "⚠ 読取失敗", C_UNMATCH, C_NOBC
            elif not mp:
                verdict, vcolor, conf_color = "❌ 未ペア", C_UNMATCH, C_NOBC
            elif conf < 0.60:
                verdict, vcolor, conf_color = "🔴 照合票要確認", C_UNMATCH, C_UNMATCH
            elif conf < 0.85 or n_unmatch > 0:
                verdict, vcolor, conf_color = "⚠ 要確認", C_REVIEW, C_REVIEW
            else:
                verdict, vcolor, conf_color = "✓ OK", C_OK, C_OK

            svals = [grp_disp, sdate, pair_disp, conf,
                     cnt[bm.STATUS_OK], n_review,
                     cnt[bm.STATUS_NO_BARCODE], n_unmatch,
                     sum(cnt.values()), verdict]
            for ci, v in enumerate(svals, 1):
                c = ws_sum.cell(sum_row, ci, v)
                c.font = base_font
                c.border = border
                if ci == 2 and isinstance(sdate, datetime):
                    c.number_format = "YYYY/MM/DD"
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif ci == 4:
                    c.number_format = "0%"
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif ci >= 5:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")
            # ระบายสี: 一致度(col4) ตามเกณฑ์, ช่อง ⚠/❌, และ 判定(col10)
            ws_sum.cell(sum_row, 4).fill = PatternFill("solid", fgColor=conf_color)
            if n_review > 0:
                ws_sum.cell(sum_row, 6).fill = PatternFill("solid", fgColor=C_REVIEW)
            if n_unmatch > 0:
                ws_sum.cell(sum_row, 8).fill = PatternFill("solid", fgColor=C_UNMATCH)
            vc = ws_sum.cell(sum_row, 10)
            vc.fill = PatternFill("solid", fgColor=vcolor)
            vc.font = Font(name="MS Gothic", size=10, bold=True, color="1F2937")
            sum_row += 1
            log_fn(f"  ✓ {grp_disp}: ✅{cnt[bm.STATUS_OK]} ⚠{cnt[bm.STATUS_REVIEW]} "
                   f"barcodeなし{cnt[bm.STATUS_NO_BARCODE]} ❌{cnt[bm.STATUS_UNMATCHED]}\n")
            success += 1
        except Exception as e:
            failed += 1
            log_fn(f"  ✗ {grp.get('no') or grp.get('key')}: {e}\n")

    # ── Phase 5: reconcile (barcode ที่เหลือไม่ถูกใช้) ──────────
    ws_rec = wb.create_sheet("照合残り")
    style_header(ws_rec, ["明細票ファイル", "未使用 品番", "未使用 Barcode"])
    rec_row = 2
    for mp, midx in meisai_indexes.items():
        for left in bm.reconcile(midx):
            ws_rec.cell(rec_row, 1, os.path.basename(mp)).font = base_font
            ws_rec.cell(rec_row, 2, left["part_no"]).font = base_font
            ws_rec.cell(rec_row, 3, left["barcode"]).font = base_font
            for ci in (1, 2, 3):
                ws_rec.cell(rec_row, ci).border = border
            rec_row += 1
    if rec_row == 2:
        ws_rec.cell(2, 1, "（未使用 barcode なし）").font = base_font

    # ── Auto width + zebra ────────────────────────────────────
    for ws, ncol in [(ws_sum, len(SUM_COLS)), (ws_det, len(DET_COLS)), (ws_rev, 9), (ws_rec, 3)]:
        for c in range(1, ncol + 1):
            mx = len(str(ws.cell(1, c).value or ""))
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, c).value
                if v is not None:
                    ln = sum(2 if ord(ch) > 127 else 1 for ch in str(v))
                    mx = max(mx, ln)
            ws.column_dimensions[get_column_letter(c)].width = max(8, min(mx * 1.1 + 2, 45))
        ws.freeze_panes = "A2"

    wb.save(out_xlsx)

    total_rows = line_no - 1
    total_bc = sum(idx["bc_total"] for idx in meisai_indexes.values())
    failed += meisai_fail + read_fail

    log_fn(f"\n{'='*60}\n✓ บันทึก: {os.path.basename(out_xlsx)}\n"
           f"  明細 {total_rows} บรรทัด · ต้องตรวจ {rev_row-2} บรรทัด\n")
    if meisai_fail or read_fail:
        log_fn(f"  ⚠⚠ 読取失敗: 明細票 {meisai_fail} / 納品書 {read_fail} ไฟล์ "
               f"(เน็ต/Azure มีปัญหาชั่วคราว — ลองรันใหม่เฉพาะไฟล์ที่ fail)\n")
    if total_bc == 0 and meisai_paths:
        log_fn("  ⚠⚠ ไม่พบ barcode เลยสักตัว — เช็คว่า: (1) มาร์กไฟล์ 注文明細票 ถูกไหม "
               "(2) Azure tier รองรับ barcode add-on ไหม\n")
    if total_rows == 0:
        log_fn("  ⚠⚠ ได้ 0 บรรทัด — น่าจะอ่านไฟล์ไม่สำเร็จทั้งหมด ดู error ด้านบนใน Log\n")
    log_fn(f"{'='*60}\n\n")
    if progress_callback:
        progress_callback(1.0)
    return success, failed


def _extract_doc_header(result, jp_helper):
    """ดึง header info จาก prebuilt-layout result"""
    info = {}

    # 1. ดึงจาก key_value_pairs (ถ้ามี)
    kv_pairs = getattr(result, "key_value_pairs", None) or []
    for kv in kv_pairs:
        if kv.key and kv.value:
            k = jp_helper.normalize_text(kv.key.content or "")
            v = jp_helper.normalize_text(kv.value.content or "")
            if k and v:
                info[k] = v

    # 2. ดึงจาก small tables (col <= 2) เช่น Table 0 ที่มี '発注日' | '2026/02/19'
    tables = getattr(result, "tables", None) or []
    for tbl in tables:
        if tbl.column_count > 2:  # ตารางใหญ่ข้ามไป เอาแค่ header table
            continue
        n_rows, n_cols = tbl.row_count, tbl.column_count
        grid = [[""] * n_cols for _ in range(n_rows)]
        for tcell in tbl.cells:
            r, c = tcell.row_index, tcell.column_index
            if r < n_rows and c < n_cols:
                grid[r][c] = jp_helper.normalize_text(tcell.content or "")
        for row in grid:
            if len(row) >= 2 and row[0].strip() and row[1].strip():
                info[row[0].strip()] = row[1].strip()

    return info


def _collect_doc_tables(result, jp_helper):
    """รวบรวม table rows จากทุกหน้า คืน (header_cols, list_of_rows)"""
    tables = getattr(result, "tables", None) or []

    def table_sort_key(t):
        if t.cells:
            br = getattr(t.cells[0], "bounding_regions", None)
            if br:
                return (br[0].page_number, t.cells[0].row_index)
        return (0, 0)

    sorted_tables = sorted(tables, key=table_sort_key)
    header_cols = None
    all_rows = []

    for tbl in sorted_tables:
        if tbl.column_count < 3:
            continue
        n_rows, n_cols = tbl.row_count, tbl.column_count
        grid = [[""] * n_cols for _ in range(n_rows)]
        for tcell in tbl.cells:
            r, c = tcell.row_index, tcell.column_index
            rs = max(getattr(tcell, "row_span", 1) or 1, 1)
            cs = max(getattr(tcell, "column_span", 1) or 1, 1)
            val = jp_helper.normalize_keep_newline(tcell.content or "")
            for dr in range(rs):
                for dc in range(cs):
                    rr, cc = r+dr, c+dc
                    if rr < n_rows and cc < n_cols:
                        grid[rr][cc] = val if (dr == 0 and dc == 0) else grid[rr][cc]

        hdr_idx = 0
        for ri in range(min(3, n_rows)):
            if any(k in " ".join(grid[ri]) for k in ["品名", "数量", "単価", "部品", "発注", "納期", "金額"]):
                hdr_idx = ri
                break

        # หา sub-header row (row ถัดจาก hdr ที่ col=0 ว่าง)
        hdr_last = hdr_idx
        if hdr_idx + 1 < n_rows and not grid[hdr_idx + 1][0].strip():
            hdr_last = hdr_idx + 1
        hdr_rows_set = set(range(hdr_idx, hdr_last + 1))

        if header_cols is None:
            header_cols = grid[hdr_idx]

        skip_rows = set()
        for ri in range(n_rows):
            if ri in skip_rows or ri in hdr_rows_set:
                continue

            current_row = list(grid[ri])

            # col=0 ว่าง = sub-row ที่ถูก skip แล้ว ข้ามไป
            if not current_row[0].strip():
                continue

            # Merge sub-rows ถัดไปที่ col=0 ว่าง
            rj = ri + 1
            while rj < n_rows and rj not in skip_rows and rj not in hdr_rows_set:
                next_row = grid[rj]
                if next_row[0].strip():  # col=0 มีค่า = data row ใหม่
                    break
                # Merge ทุก col ที่ sub-row มีค่า
                for c in range(n_cols):
                    nv = next_row[c].strip()
                    cv = current_row[c].strip()
                    if nv:
                        if not cv:
                            # col ว่าง → เติมค่าจาก sub-row
                            current_row[c] = nv
                        else:
                            # col มีค่าแล้ว → ต่อท้ายด้วย \n เพื่อเก็บทั้งคู่
                            current_row[c] = cv + "\n" + nv
                skip_rows.add(rj)
                rj += 1

            row_text = " ".join(current_row)
            if not row_text.strip():
                continue
            if any(k in row_text for k in ["小計", "合計", "総計"]):
                continue
            all_rows.append(current_row)

    return header_cols or [], all_rows


def render_table_faithful(ws, table, jp_helper, start_row=1):
    """
    เขียนตาราง Azure ลง worksheet "ตามจริง" — ทุกแถว/คอลัมน์ + merge cell
    ความกว้างคอลัมน์/ความสูงแถว = ตามพิกัด bounding box จริง (สัดส่วนใกล้ต้นฉบับ)
    เก็บข้อความดิบ ไม่แปลงค่า กันข้อมูลเพี้ยน. คืน row ถัดไป (ไว้ stack หลายตาราง)
    """
    n_rows = getattr(table, "row_count", 0)
    n_cols = getattr(table, "column_count", 0)
    if n_rows <= 0 or n_cols <= 0:
        return start_row

    grid = [["" for _ in range(n_cols)] for _ in range(n_rows)]
    spans = []
    col_l, col_r, row_t, row_b = {}, {}, {}, {}   # ขอบเขตจริง (หน่วยนิ้ว) ต่อคอลัมน์/แถว
    for cell in table.cells:
        r, c = cell.row_index, cell.column_index
        rs = max(getattr(cell, "row_span", 1) or 1, 1)
        cs = max(getattr(cell, "column_span", 1) or 1, 1)
        if r < n_rows and c < n_cols:
            grid[r][c] = jp_helper.normalize_keep_newline(cell.content or "")
        if rs > 1 or cs > 1:
            spans.append((r, c, rs, cs))
        brs = getattr(cell, "bounding_regions", None)
        if brs:
            poly = brs[0].polygon
            xs, ys = poly[0::2], poly[1::2]
            if xs and ys:
                x0, x1, y0, y1 = min(xs), max(xs), min(ys), max(ys)
                if cs == 1:   # ใช้เฉพาะช่องไม่ merge เพื่อหาขอบคอลัมน์/แถวที่สะอาด
                    col_l[c] = min(col_l.get(c, x0), x0)
                    col_r[c] = max(col_r.get(c, x1), x1)
                if rs == 1:
                    row_t[r] = min(row_t.get(r, y0), y0)
                    row_b[r] = max(row_b.get(r, y1), y1)

    thin = Side(style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in range(n_rows):
        for c in range(n_cols):
            cell = ws.cell(start_row + r, c + 1, grid[r][c])
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(name="MS Gothic", size=10,
                             bold=(r == 0))   # แถวแรก = หัวตาราง (ตัวหนาเฉย ๆ ไม่ใส่สี เพื่อใกล้ต้นฉบับ)

    for (r, c, rs, cs) in spans:
        try:
            ws.merge_cells(start_row=start_row + r, start_column=c + 1,
                           end_row=start_row + min(r + rs - 1, n_rows - 1),
                           end_column=c + 1 + min(cs - 1, n_cols - 1 - c))
        except Exception:
            pass

    # ── ความกว้างคอลัมน์ตามพิกัดจริง (นิ้ว → หน่วยความกว้าง Excel) ──
    widths = {c: (col_r[c] - col_l[c]) for c in col_l if col_r.get(c, 0) > col_l[c]}
    if widths:
        avg_w = sum(widths.values()) / len(widths)
        for c in range(n_cols):
            w_in = widths.get(c, avg_w)
            w_chars = max(3.0, min(w_in * 13.7 - 0.7, 120))   # px=inch*96, width≈(px-5)/7
            col = get_column_letter(c + 1)
            ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 0, w_chars)
    else:  # ไม่มีพิกัด → fallback ตามความยาวข้อความ
        for c in range(n_cols):
            mx = max((max((sum(2 if ord(ch) > 127 else 1 for ch in ln) for ln in str(grid[r][c]).split("\n")), default=0)
                      for r in range(n_rows)), default=6)
            col = get_column_letter(c + 1)
            ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 0, min(mx * 1.1 + 2, 60))

    # ── ความสูงแถวตามพิกัดจริง (นิ้ว → point) ──
    heights = {r: (row_b[r] - row_t[r]) for r in row_t if row_b.get(r, 0) > row_t[r]}
    if heights:
        avg_h = sum(heights.values()) / len(heights)
        for r in range(n_rows):
            h_in = heights.get(r, avg_h)
            ws.row_dimensions[start_row + r].height = max(13.0, min(h_in * 72, 400))

    return start_row + n_rows


def _safe_sheet_name(base, used):
    """ชื่อ sheet Excel: ตัดอักขระต้องห้าม + ≤31 ตัว + ไม่ซ้ำ"""
    s = re.sub(r'[\[\]:\*\?/\\]', "_", str(base)).strip() or "Sheet"
    s = s[:28]
    name = s
    i = 1
    while name in used:
        name = f"{s[:25]}_{i}"
        i += 1
    used.add(name)
    return name


def export_documents_to_excel(pdf_paths: list, out_xlsx: str, log_fn, progress_callback=None, combine=False):
    """
    Document mode = ตัวแปลง PDF → Excel ทั่วไป (prebuilt-layout) แบบ "ตามจริง"
      combine=False : 1 ตาราง = 1 sheet (ชื่อ sheet ตามไฟล์)
      combine=True  : รวมทุกตารางไว้ใน sheet เดียว (คั่นด้วยชื่อไฟล์)
    ดึงทุกตาราง ไม่กรองด้วยกฎธุรกิจ ไม่ตัดแถว เก็บข้อความดิบ
    """
    jp_helper = JapaneseHelper()
    log_fn(f"\n{'='*60}\nDocument mode (PDF→Excel ตามจริง): {len(pdf_paths)} files"
           f" — {'รวมชีตเดียว' if combine else '1ไฟล์=1ชีต'}\n{'='*60}\n\n")

    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()
    success = 0
    failed = 0
    n = max(1, len(pdf_paths))

    ws_comb = None
    comb_row = 1
    if combine:
        ws_comb = wb.create_sheet("All_Tables")

    for idx, pdf_path in enumerate(pdf_paths, 1):
        fname = os.path.basename(pdf_path)
        log_fn(f"[{idx}/{len(pdf_paths)}] {fname}\n")
        if progress_callback:
            progress_callback((idx - 1) / n * 0.95)
        try:
            result = analyze_document(pdf_path, log_fn=log_fn)
            tables = getattr(result, "tables", None) or []
            if not tables:
                log_fn("  ⚠ ไม่พบตารางในไฟล์นี้\n")
                if combine:
                    ws_comb.cell(comb_row, 1, f"【{fname}】 — テーブルなし").font = Font(bold=True, color="B71C1C")
                    comb_row += 2
                success += 1
                continue

            if combine:
                ws_comb.cell(comb_row, 1, f"【{fname}】").font = Font(name="MS Gothic", size=12, bold=True, color="1F4E79")
                comb_row += 1
                for t in tables:
                    comb_row = render_table_faithful(ws_comb, t, jp_helper, start_row=comb_row)
                    comb_row += 2  # เว้นบรรทัดคั่นตาราง
            else:
                multi = len(tables) > 1
                for ti, t in enumerate(tables, 1):
                    base = os.path.splitext(fname)[0] + (f"_T{ti}" if multi else "")
                    ws = wb.create_sheet(_safe_sheet_name(base, used_names))
                    render_table_faithful(ws, t, jp_helper, start_row=1)
                    ws.freeze_panes = "A2"

            log_fn(f"  ✓ {len(tables)} ตาราง\n")
            success += 1
        except Exception as e:
            failed += 1
            log_fn(f"  ✗✗ 読取失敗: {type(e).__name__}: {e}\n")

    if not wb.worksheets:
        wb.create_sheet("Empty")
    wb.save(out_xlsx)
    if progress_callback:
        progress_callback(1.0)
    log_fn(f"\n{'='*60}\n✓ Saved: {os.path.basename(out_xlsx)}\n"
           f"  Success: {success}, Failed: {failed}\n{'='*60}\n\n")
    return success, failed

def process_invoice_with_progress(pdf_path: str, out_xlsx: str, log_fn,
                                  progress_callback=None, file_idx=0, total_files=1):
    """Process invoice with detailed progress updates"""

    base_progress = file_idx / total_files
    file_progress_size = 1.0 / total_files

    def update_file_progress(step_progress):
        """Update progress for current file (0.0 to 1.0)"""
        if progress_callback:
            total_progress = base_progress + (step_progress * file_progress_size)
            progress_callback(total_progress)

    log_fn(f"Processing: {os.path.basename(pdf_path)}")

    # Step 1: Analyzing (0% → 70%)
    update_file_progress(0.0)
    simulator = ProgressSimulator(update_file_progress, 0.0, 0.7, duration=2.0)
    simulator.start_simulation()

    try:
        result = analyze_invoice(pdf_path)
        simulator.stop_simulation()
        update_file_progress(0.7)
    except Exception as e:
        simulator.stop_simulation()
        raise e

    # Step 2: Extracting (70% → 85%)
    update_file_progress(0.7)
    jp_helper = JapaneseHelper()
    header, items = extract_fields(result, jp_helper)
    update_file_progress(0.85)

    # Step 3: Creating Excel (85% → 100%)
    update_file_progress(0.85)
    create_beautiful_excel(header, pdf_path, out_xlsx, log_fn, result)
    update_file_progress(1.0)

# ======================
# UI
# ======================
class InvoiceApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("KSJ Invoice Pro")
        self.geometry("1060x620")
        self.minsize(1000, 580)
        self.resizable(True, True)
        if os.path.exists(icon_path):
            self.iconbitmap(icon_path)


        self.files = []
        self.meisai_files = set()
        self._file_btns = []
        self.log_visible = False
        self.task = "invoice"

        self.create_widgets()

    # ── Modern Light palette ──
    BG = "#F4F6F9"
    CARD = "#FFFFFF"
    ACCENT = "#2563EB"
    ACCENT_DK = "#1D4ED8"
    ACCENT_LT = "#DBEAFE"
    TEXT = "#1F2937"
    MUTED = "#6B7280"
    BORDER = "#E5E7EB"
    SUCCESS = "#16A34A"
    SUCCESS_DK = "#15803D"
    AMBER = "#D97706"
    AMBER_BG = "#FEF3C7"

    # ── Task definitions (sidebar) ──────────────────────────────
    #  key, sidebar label (short), main title, description
    TASK_DEFS = [
        ("invoice",  "📊  請求書・納品書",      "請求書・納品書 → Excel",
         "スキャンPDFを読み取って表形式のExcelに変換します"),
        ("barcode",  "🏷  バーコード照合",      "バーコード照合（注文明細票 ↔ 納品書）",
         "注文明細票のバーコードを納品書の各行に照合します"),
        ("document", "📄  書類をそのまま変換",   "書類 → Excel（そのまま）",
         "レイアウトを保持して書類をExcelに変換します"),
        ("po",       "🧾  発注書 → CSV",        "発注書 → CSV",
         "仕入先ごとの注文書を読み取り、取込用ファイルに変換します"),
    ]

    def create_widgets(self):
        from tkinter import ttk

        ctk.set_appearance_mode("light")
        self.configure(fg_color=self.BG)

        self._task_meta = {k: (title, desc) for k, _s, title, desc in self.TASK_DEFS}

        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1, minsize=320)

        # ── Header ──────────────────────────────────────────────
        header_frame = ctk.CTkFrame(self, fg_color=self.ACCENT, corner_radius=0, height=58)
        header_frame.grid(row=0, column=0, columnspan=2, sticky="ew")
        header_frame.grid_propagate(False)

        title_box = ctk.CTkFrame(header_frame, fg_color="transparent")
        title_box.pack(side='left', padx=20, pady=8)
        ctk.CTkLabel(title_box, text="KSJ Invoice Pro",
                     font=ctk.CTkFont(size=20, weight="bold"),
                     text_color="white").pack(anchor='w')
        ctk.CTkLabel(title_box, text="PDF 変換ツール",
                     font=ctk.CTkFont(size=11),
                     text_color="#C7DCFF").pack(anchor='w')

        # ── Sidebar (task selector) ─────────────────────────────
        sidebar = ctk.CTkFrame(self, fg_color=self.CARD, corner_radius=0, width=216)
        sidebar.grid(row=1, column=0, sticky="nsew")
        sidebar.grid_propagate(False)

        ctk.CTkLabel(sidebar, text="タスクを選択",
                     font=ctk.CTkFont(size=11),
                     text_color=self.MUTED).pack(anchor='w', padx=16, pady=(14, 6))

        self._task_btns = {}
        for key, short, _title, _desc in self.TASK_DEFS:
            b = ctk.CTkButton(
                sidebar, text=short, anchor="w",
                height=44, corner_radius=8,
                font=ctk.CTkFont(size=12),
                fg_color="transparent", text_color=self.TEXT,
                hover_color=self.ACCENT_LT,
                command=lambda k=key: self.select_task(k)
            )
            b.pack(fill='x', padx=10, pady=2)
            self._task_btns[key] = b

        # ── Main ────────────────────────────────────────────────
        main_frame = ctk.CTkFrame(self, fg_color=self.BG, corner_radius=0)
        main_frame.grid(row=1, column=1, sticky="nsew", padx=0, pady=0)
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_rowconfigure(3, weight=1)

        # ── Task header (title + description) ────────────────────
        task_head = ctk.CTkFrame(main_frame, fg_color="transparent")
        task_head.grid(row=0, column=0, sticky="ew", padx=16, pady=(14, 2))
        self.task_title = ctk.CTkLabel(task_head, text="",
                                       font=ctk.CTkFont(size=16, weight="bold"),
                                       text_color=self.TEXT)
        self.task_title.pack(anchor='w')
        self.task_desc = ctk.CTkLabel(task_head, text="",
                                      font=ctk.CTkFont(size=11),
                                      text_color=self.MUTED)
        self.task_desc.pack(anchor='w')

        # ── Options panel (per-task) ────────────────────────────
        self.options_panel = ctk.CTkFrame(main_frame, fg_color=self.CARD, corner_radius=10)
        self.options_panel.grid(row=1, column=0, sticky="ew", padx=16, pady=(8, 8))
        self._opt_row = ctk.CTkFrame(self.options_panel, fg_color="transparent")
        self._opt_row.pack(fill='x', padx=14, pady=10)

        # PO vendor + format selectors (shown only for 発注書→CSV)
        self.po_vendor = ctk.StringVar(value=list(poc.VENDORS.keys())[0])
        self.po_vendor_label = ctk.CTkLabel(self._opt_row, text="仕入先",
                                            font=ctk.CTkFont(size=11, weight="bold"),
                                            text_color=self.TEXT)
        self.po_vendor_dropdown = ctk.CTkOptionMenu(
            self._opt_row, variable=self.po_vendor,
            values=list(poc.VENDORS.keys()),
            width=190, height=30, corner_radius=8,
            font=ctk.CTkFont(size=11),
            fg_color="#0E7490", button_color="#155E75",
            button_hover_color="#0C4A6E", text_color="white",
        )

        self.po_fmt_map = {"Excelのみ": "xlsx", "CSV + Excel": "both", "CSVのみ": "csv"}
        self.po_format = ctk.StringVar(value="Excelのみ")
        self.po_format_label = ctk.CTkLabel(self._opt_row, text="出力形式",
                                            font=ctk.CTkFont(size=11, weight="bold"),
                                            text_color=self.TEXT)
        self.po_format_dropdown = ctk.CTkOptionMenu(
            self._opt_row, variable=self.po_format,
            values=list(self.po_fmt_map.keys()),
            width=130, height=30, corner_radius=8,
            font=ctk.CTkFont(size=11),
            fg_color="#0E7490", button_color="#155E75",
            button_hover_color="#0C4A6E", text_color="white",
        )

        # Combine checkbox (label changes per task)
        self.combine_mode = ctk.BooleanVar(value=False)
        self.combine_checkbox = ctk.CTkCheckBox(
            self._opt_row, text="全PDFを1つのExcelにまとめる",
            variable=self.combine_mode,
            font=ctk.CTkFont(size=11), text_color=self.TEXT,
            fg_color=self.ACCENT, hover_color=self.ACCENT_DK,
            checkmark_color="white", corner_radius=5
        )

        # Barcode mode flag (driven by task, not a visible checkbox anymore)
        self.barcode_mode = ctk.BooleanVar(value=False)
        self.barcode_note = ctk.CTkLabel(
            self._opt_row,
            text="📋 注文明細票 のファイルをリストでダブルクリックしてマークしてください",
            font=ctk.CTkFont(size=11), text_color=self.AMBER)

        self.combine_mode.trace_add("write", self._on_combine_toggle)

        # ── File table toolbar ───────────────────────────────────
        toolbar = ctk.CTkFrame(main_frame, fg_color=self.CARD, corner_radius=0, height=44)
        toolbar.grid(row=2, column=0, sticky="ew", padx=12, pady=(0, 0))
        toolbar.grid_propagate(False)

        btn_cfg = dict(width=32, height=30, corner_radius=8,
                       fg_color=self.ACCENT, hover_color=self.ACCENT_DK,
                       text_color="white", font=ctk.CTkFont(size=14, weight="bold"))

        ctk.CTkButton(toolbar, text="＋", **btn_cfg,
                      command=self.pick_files).pack(side='left', padx=(8, 2), pady=6)
        ctk.CTkButton(toolbar, text="－", **btn_cfg,
                      command=self.remove_selected).pack(side='left', padx=2, pady=6)
        ctk.CTkButton(toolbar, text="▲", **btn_cfg,
                      command=self.move_up).pack(side='left', padx=(10, 2), pady=6)
        ctk.CTkButton(toolbar, text="▼", **btn_cfg,
                      command=self.move_down).pack(side='left', padx=2, pady=6)

        ctk.CTkButton(toolbar, text="名前順 A→Z", width=84, height=30, corner_radius=8,
                      fg_color="#EFF3FA", hover_color=self.ACCENT_LT,
                      text_color=self.ACCENT, font=ctk.CTkFont(size=11),
                      command=self.sort_files).pack(side='left', padx=(10, 2), pady=6)

        ctk.CTkButton(toolbar, text="🗑 全削除", width=84, height=30, corner_radius=8,
                      fg_color="#EF4444", hover_color="#DC2626",
                      text_color="white", font=ctk.CTkFont(size=11),
                      command=self.clear_files).pack(side='left', padx=2, pady=6)

        self.file_count = ctk.CTkLabel(toolbar, text="0 ファイル",
                                       font=ctk.CTkFont(size=11),
                                       text_color=self.MUTED)
        self.file_count.pack(side='left', padx=12)

        self.meisai_hint = ctk.CTkLabel(toolbar, text="",
                                        font=ctk.CTkFont(size=11, weight="bold"),
                                        text_color=self.AMBER)
        self.meisai_hint.pack(side='left', padx=5)

        # ── File table (ttk.Treeview) ───────────────────────────
        table_frame = ctk.CTkFrame(main_frame, fg_color=self.CARD, corner_radius=0)
        table_frame.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 0))
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("File.Treeview",
                        background=self.CARD, foreground=self.TEXT,
                        fieldbackground=self.CARD, borderwidth=0,
                        rowheight=27, font=("Yu Gothic UI", 10))
        style.configure("File.Treeview.Heading",
                        background=self.ACCENT, foreground="white",
                        font=("Yu Gothic UI", 10, "bold"), relief="flat", padding=4)
        style.map("File.Treeview.Heading", background=[("active", self.ACCENT_DK)])
        style.map("File.Treeview",
                  background=[("selected", self.ACCENT_LT)],
                  foreground=[("selected", self.TEXT)])

        cols = ("状態", "フォルダー名", "ファイル名", "サイズ", "更新日時")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings",
                                 style="File.Treeview", selectmode="extended")

        col_widths = {"状態": 110, "フォルダー名": 130, "ファイル名": 280, "サイズ": 75, "更新日時": 130}
        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=col_widths[col], minwidth=40,
                             anchor="center" if col in ("状態", "サイズ", "更新日時") else "w")

        # 注文明細票 = amber chip; แถวปกติ zebra
        self.tree.tag_configure("meisai", background=self.AMBER_BG, foreground="#92400E")
        self.tree.tag_configure("normal", background=self.CARD, foreground=self.TEXT)
        self.tree.tag_configure("normal_alt", background="#F8FAFC", foreground=self.TEXT)

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        self.tree.bind("<Double-1>", self._on_tree_double_click)

        try:
            self.drop_target_register('DND_Files')
            self.dnd_bind('<<Drop>>', self._on_drop)
        except Exception:
            pass

        # ── Bottom bar ──────────────────────────────────────────
        bottom = ctk.CTkFrame(main_frame, fg_color=self.CARD, corner_radius=0)
        bottom.grid(row=4, column=0, sticky="ew", padx=12, pady=(0, 12))
        bottom.grid_columnconfigure(1, weight=1)

        # Output row
        out_row = ctk.CTkFrame(bottom, fg_color="transparent")
        out_row.pack(fill='x', padx=12, pady=(12, 6))

        ctk.CTkLabel(out_row, text="出力先",
                     font=ctk.CTkFont(size=11, weight="bold"),
                     text_color=self.TEXT).pack(side='left')

        self.out_dir = ctk.StringVar(value=DEFAULT_OUT_DIR)
        ctk.CTkEntry(out_row, textvariable=self.out_dir,
                     font=ctk.CTkFont(family="Consolas", size=10),
                     fg_color="#FFFFFF", text_color=self.TEXT,
                     border_color=self.BORDER, border_width=1, height=30,
                     corner_radius=8
                     ).pack(side='left', fill='x', expand=True, padx=(10, 6))

        ctk.CTkButton(out_row, text="参照…", width=70, height=30, corner_radius=8,
                      fg_color="#EFF3FA", hover_color=self.ACCENT_LT,
                      text_color=self.ACCENT, font=ctk.CTkFont(size=11),
                      command=self.pick_dir).pack(side='left')

        # Action row
        action_row = ctk.CTkFrame(bottom, fg_color="transparent")
        action_row.pack(fill='x', padx=12, pady=(0, 12))

        self.run_btn = ctk.CTkButton(action_row, text="▶  実行",
                                     command=self.run_batch,
                                     width=120, height=40, corner_radius=10,
                                     font=ctk.CTkFont(size=14, weight="bold"),
                                     fg_color=self.SUCCESS, hover_color=self.SUCCESS_DK)
        self.run_btn.pack(side='left')

        # Status + progress
        status_mid = ctk.CTkFrame(action_row, fg_color="transparent")
        status_mid.pack(side='left', fill='x', expand=True, padx=14)

        status_top = ctk.CTkFrame(status_mid, fg_color="transparent")
        status_top.pack(fill='x')

        self.status_label = ctk.CTkLabel(status_top, text="● Ready",
                                         font=ctk.CTkFont(size=11),
                                         text_color=self.SUCCESS)
        self.status_label.pack(side='left')

        self.progress_pct_label = ctk.CTkLabel(status_top, text="0%",
                                               font=ctk.CTkFont(size=11, weight="bold"),
                                               text_color=self.ACCENT)
        self.progress_pct_label.pack(side='left', padx=(10, 0))

        self.toggle_btn = ctk.CTkButton(status_top, text="▼ Log",
                                        width=70, height=24, corner_radius=8,
                                        fg_color="#EFF3FA", hover_color=self.ACCENT_LT,
                                        text_color=self.ACCENT,
                                        font=ctk.CTkFont(size=10),
                                        command=self.toggle_log)
        self.toggle_btn.pack(side='right')

        self.progress_bar = ctk.CTkProgressBar(status_mid, height=8, corner_radius=4,
                                               progress_color=self.ACCENT,
                                               fg_color="#E8EDF5")
        self.progress_bar.pack(fill='x', pady=(6, 0))
        self.progress_bar.set(0)

        # Log (hidden)
        self.log_frame = ctk.CTkFrame(main_frame, fg_color="#0F172A", corner_radius=0, height=170)
        self.log_frame.grid(row=5, column=0, sticky="ew", padx=0, pady=0)
        self.log_frame.grid_propagate(False)
        self.log_frame.grid_remove()

        self.log = ctk.CTkTextbox(self.log_frame, height=150,
                                  font=ctk.CTkFont(family="Consolas", size=10),
                                  fg_color="#0F172A", text_color="#E2E8F0")
        self.log.pack(fill='both', expand=True, padx=10, pady=8)

        self.log.insert("end", "KSJ Invoice Pro — Ready\n")

        # mode var เก็บไว้ให้ run_batch / helper เดิมใช้ (ขับด้วย select_task)
        self.doc_mode = ctk.StringVar(value="invoice")

        # เลือกงานเริ่มต้น
        self.select_task("invoice")

    def _layout_options(self, name):
        """โชว์เฉพาะ option ของงานที่เลือก (ตัวอื่น pack_forget)"""
        for w in (self.po_vendor_label, self.po_vendor_dropdown,
                  self.po_format_label, self.po_format_dropdown,
                  self.combine_checkbox, self.barcode_note):
            w.pack_forget()

        if name == "po":
            self.po_vendor_label.pack(side='left', padx=(0, 6))
            self.po_vendor_dropdown.pack(side='left', padx=(0, 16))
            self.po_format_label.pack(side='left', padx=(0, 6))
            self.po_format_dropdown.pack(side='left', padx=(0, 16))
            self.combine_checkbox.configure(text="全PDFを1つにまとめる")
            self.combine_checkbox.pack(side='left')
        elif name == "invoice":
            self.combine_checkbox.configure(text="全PDFを1つのExcelにまとめる")
            self.combine_checkbox.pack(side='left')
        elif name == "document":
            self.combine_checkbox.configure(text="全テーブルを1シートにまとめる")
            self.combine_checkbox.pack(side='left')
        elif name == "barcode":
            self.barcode_note.pack(side='left')

    def select_task(self, name):
        """สลับงาน: ไฮไลต์ sidebar + ตั้งตัวแปร mode เดิมให้ run_batch ใช้"""
        self.task = name

        for key, btn in self._task_btns.items():
            if key == name:
                btn.configure(fg_color=self.ACCENT, text_color="white",
                              hover_color=self.ACCENT_DK)
            else:
                btn.configure(fg_color="transparent", text_color=self.TEXT,
                              hover_color=self.ACCENT_LT)

        if name == "invoice":
            self.doc_mode.set("invoice")
            self.barcode_mode.set(False)
        elif name == "barcode":
            self.doc_mode.set("invoice")
            self.barcode_mode.set(True)
            self.combine_mode.set(True)   # barcode照合 = combine-style export
        elif name == "document":
            self.doc_mode.set("document")
            self.barcode_mode.set(False)
        elif name == "po":
            self.doc_mode.set("発注書→CSV")
            self.barcode_mode.set(False)

        title, desc = self._task_meta[name]
        self.task_title.configure(text=title)
        self.task_desc.configure(text=desc)

        self._layout_options(name)
        self._refresh_tree()

    def _on_combine_toggle(self, *args):
        self._refresh_tree()

    def toggle_log(self):
        if self.log_visible:
            self.log_frame.grid_remove()
            self.toggle_btn.configure(text="▼ Log")
            self.log_visible = False
        else:
            self.log_frame.grid()
            self.toggle_btn.configure(text="▲ Log")
            self.log_visible = True

    def pick_files(self):
        paths = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
        if paths:
            existing = set(self.files)
            for p in paths:
                if p not in existing:
                    self.files.append(p)
            self.update_file_list()

    def clear_files(self):
        if not self.files:
            return
        if not messagebox.askyesno("確認", f"ファイル {len(self.files)} 件をすべて削除しますか？"):
            return
        self.files = []
        self.meisai_files = set()
        self.update_file_list()

    def sort_files(self):
        """เรียงไฟล์ตามชื่อ A→Z (ช่วยให้ใบหลายหน้า _0019/_0020 เรียงถูกลำดับ)"""
        self.files.sort(key=lambda p: os.path.basename(p).lower())
        self.update_file_list()

    def remove_selected(self):
        selected = self.tree.selection()
        if not selected:
            return
        indices = sorted([self.tree.index(iid) for iid in selected], reverse=True)
        for i in indices:
            path = self.files[i]
            self.meisai_files.discard(path)
            self.files.pop(i)
        self.update_file_list()

    def move_up(self):
        selected = self.tree.selection()
        if not selected:
            return
        idx = self.tree.index(selected[0])
        if idx > 0:
            self.files[idx-1], self.files[idx] = self.files[idx], self.files[idx-1]
            self.update_file_list()
            # re-select
            items = self.tree.get_children()
            if idx-1 < len(items):
                self.tree.selection_set(items[idx-1])

    def move_down(self):
        selected = self.tree.selection()
        if not selected:
            return
        idx = self.tree.index(selected[0])
        if idx < len(self.files) - 1:
            self.files[idx], self.files[idx+1] = self.files[idx+1], self.files[idx]
            self.update_file_list()
            items = self.tree.get_children()
            if idx+1 < len(items):
                self.tree.selection_set(items[idx+1])

    def _on_tree_double_click(self, event):
        iid = self.tree.identify_row(event.y)
        if not iid:
            return
        idx = self.tree.index(iid)
        path = self.files[idx]
        if path in self.meisai_files:
            self.meisai_files.discard(path)
        else:
            self.meisai_files.add(path)
        self._refresh_tree()

    def _on_drop(self, event):
        paths = self.tk.splitlist(event.data)
        existing = set(self.files)
        for p in paths:
            if p.lower().endswith(".pdf") and p not in existing:
                self.files.append(p)
        self.update_file_list()

    def _refresh_tree(self):
        """อัพเดทสี row และ hint label"""
        for iid in self.tree.get_children():
            idx = self.tree.index(iid)
            path = self.files[idx]
            is_meisai = path in self.meisai_files
            tag = "meisai" if is_meisai else ("normal" if idx % 2 == 0 else "normal_alt")
            status = "📋 注文明細票" if is_meisai else "—"
            self.tree.item(iid, values=(
                status,
                self.tree.item(iid)["values"][1],
                self.tree.item(iid)["values"][2],
                self.tree.item(iid)["values"][3],
                self.tree.item(iid)["values"][4],
            ), tags=(tag,))

        n_meisai = len(self.meisai_files)
        n_delivery = len(self.files) - n_meisai
        if getattr(self, "task", None) == "barcode":
            if n_meisai > 0:
                self.meisai_hint.configure(
                    text=f"📋 注文明細票: {n_meisai}  |  📄 納品書: {n_delivery}   (ダブルクリック = マーク)"
                )
            else:
                self.meisai_hint.configure(
                    text="⚠ 注文明細票のファイルをダブルクリックしてマークしてください"
                )
        else:
            self.meisai_hint.configure(text="")

    def update_file_list(self):
        import os, datetime
        for iid in self.tree.get_children():
            self.tree.delete(iid)

        for i, path in enumerate(self.files):
            is_meisai = path in self.meisai_files
            tag = "meisai" if is_meisai else ("normal" if i % 2 == 0 else "normal_alt")
            status = "📋 注文明細票" if is_meisai else "—"
            folder = os.path.dirname(path)
            fname = os.path.basename(path)
            try:
                size = os.path.getsize(path)
                size_str = f"{size//1024} KB" if size < 1024*1024 else f"{size//(1024*1024)} MB"
                mtime = os.path.getmtime(path)
                dt_str = datetime.datetime.fromtimestamp(mtime).strftime("%y/%m/%d %H:%M:%S")
            except Exception:
                size_str = "—"
                dt_str = "—"
            self.tree.insert("", "end",
                             values=(status, folder, fname, size_str, dt_str),
                             tags=(tag,))

        count = len(self.files)
        self.file_count.configure(text=f"{count} ファイル")
        self._refresh_tree()
        self.append_log(f"✓ {count} file(s) selected\n")

    def pick_dir(self):
        d = filedialog.askdirectory(initialdir=self.out_dir.get() or os.getcwd())
        if d:
            self.out_dir.set(d)
            save_settings(out_dir=d)   # จำที่เลือกล่าสุด

    def append_log(self, msg: str):
        # thread-safe: Tk ไม่ปลอดภัยถ้าเรียกจาก worker thread → marshal เข้า main loop
        if threading.current_thread() is threading.main_thread():
            self._append_log_ui(msg)
        else:
            try:
                self.after(0, lambda m=msg: self._append_log_ui(m))
            except Exception:
                pass

    def _append_log_ui(self, msg: str):
        try:
            self.log.insert("end", msg)
            self.log.see("end")
        except Exception:
            pass

    def update_status(self, text: str, color: str = None):
        self.status_label.configure(text=text, text_color=color or self.SUCCESS)

    def update_progress(self, value: float):
        """Update progress bar with percentage (0.0 to 1.0)"""
        value = max(0.0, min(1.0, value))
        self.progress_bar.set(value)

        # Update percentage label
        pct = int(value * 100)
        self.progress_pct_label.configure(text=f"{pct}%")

        self.update_idletasks()

    def run_batch(self):
        if not self.files:
            messagebox.showwarning("ファイル未選択", "PDFファイルを選択してください")
            return

        out_dir = self.out_dir.get().strip()
        if not out_dir:
            messagebox.showwarning("出力先未選択", "出力フォルダを選択してください")
            return

        os.makedirs(out_dir, exist_ok=True)
        save_settings(out_dir=out_dir)   # จำที่ใช้ล่าสุด

        # เปิด Log ให้เห็นเสมอตอนรัน — กัน error เงียบ
        if not self.log_visible:
            self.toggle_log()

        self.run_btn.configure(state="disabled")
        total = len(self.files)
        is_combine = self.combine_mode.get()
        is_barcode = self.barcode_mode.get()
        print(f"DEBUG run: doc_mode={self.doc_mode.get()} is_combine={is_combine} is_barcode={is_barcode} meisai={len(self.meisai_files)}")

        # Validate: barcode mode ต้องมี user mark ไฟล์ 注文明細票 ไว้
        if is_combine and is_barcode:
            if not self.meisai_files:
                messagebox.showerror(
                    "ファイルがマークされていません",
                    "注文明細票のファイルをリストからクリックしてマークしてください\n"
                    "(マークされたファイルは茶色で表示されます 📋)"
                )
                self.run_btn.configure(state="normal")
                return

        def progress_callback(value, total=None):
            """Accept either progress fraction (0..1) or (current, total)."""
            if total is not None:
                try:
                    value = (float(value) / float(total)) if float(total) else 0.0
                except Exception:
                    value = 0.0
            self.after(0, lambda v=float(value): self.update_progress(v))

        is_document = self.doc_mode.get() == "document"
        is_po = self.doc_mode.get() == "発注書→CSV"
        po_vendor = self.po_vendor.get()
        po_fmt = self.po_fmt_map.get(self.po_format.get(), "both")

        def worker():
            ok = 0
            fail = 0

            if is_po:
                # PO→CSV mode: Azure DI + parser ต่อเจ้า → CSV ต่อไฟล์
                self.after(0, lambda: self.update_status(
                    f"● [発注書→CSV] {po_vendor} ...", "#0E7490"))
                try:
                    ok, fail = poc.export_po_to_csv(
                        self.files, out_dir, po_vendor, ENDPOINT, KEY,
                        self.append_log, progress_callback, combine=is_combine,
                        fmt=po_fmt
                    )
                    self.after(0, lambda: self.update_progress(1.0))
                except Exception as e:
                    self.append_log(f"\n✗ Error: {str(e)}\n")
                    fail = total

            elif is_document:
                # Document mode: prebuilt-layout รวมทุกไฟล์เป็น Excel เดียว
                self.after(0, lambda: self.update_status(
                    "● [Document] Combining all PDFs...", "#e67e22"))
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                out_xlsx = os.path.join(out_dir, f"COMBINED_DOCUMENTS_{ts}.xlsx")
                try:
                    ok, fail = export_documents_to_excel(
                        self.files, out_xlsx, self.append_log, progress_callback,
                        combine=is_combine
                    )
                    self.after(0, lambda: self.update_progress(1.0))
                except Exception as e:
                    self.append_log(f"\n✗ Error: {str(e)}\n")
                    fail = total

            elif is_combine and is_barcode:
                # NEW Barcode mode: meisai (layout+barcodes) ↔ delivery
                self.after(0, lambda: self.update_status(
                    "● Barcode照合 中...", "#e67e22"))

                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                out_xlsx = os.path.join(out_dir, f"BARCODE_照合_{ts}.xlsx")

                meisai_paths = [p for p in self.files if p in self.meisai_files]
                delivery_paths = [p for p in self.files if p not in self.meisai_files]
                try:
                    ok, fail = export_barcode_mode(
                        meisai_paths, delivery_paths, out_xlsx,
                        self.append_log, progress_callback
                    )
                    self.after(0, lambda: self.update_progress(1.0))
                except Exception as e:
                    self.append_log(f"\n✗ Error: {str(e)}\n")
                    fail = total

            elif is_combine:
                # Invoice Combine mode (ไม่มี barcode)
                self.after(0, lambda: self.update_status(
                    "● Combining all PDFs to ONE Excel...", "#2d89ef"))

                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                out_xlsx = os.path.join(out_dir, f"COMBINED_INVOICES_{ts}.xlsx")

                try:
                    ok, fail = export_many_to_one_excel(
                        self.files, out_xlsx, self.append_log, progress_callback,
                        barcode_mode=False,
                        meisai_override=None
                    )
                    self.after(0, lambda: self.update_progress(1.0))
                except Exception as e:
                    self.append_log(f"\n✗ Error: {str(e)}\n")
                    fail = total

            else:
                # Invoice Separate mode
                for idx, pdf_path in enumerate(self.files):
                    try:
                        self.after(0, lambda idx=idx: self.update_status(
                            f"● Processing {idx+1}/{total}: {os.path.basename(pdf_path)}",
                            "#2d89ef"
                        ))
                        base = os.path.splitext(os.path.basename(pdf_path))[0]
                        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                        out_xlsx = os.path.join(out_dir, f"{base}_{ts}.xlsx")

                        process_invoice_with_progress(
                            pdf_path, out_xlsx, self.append_log,
                            progress_callback, idx, total
                        )
                        ok += 1
                    except Exception as e:
                        self.append_log(f"\n✗ Error: {str(e)}\n")
                        fail += 1

            self.after(0, lambda: self._done(ok, fail, is_combine))

        threading.Thread(target=worker, daemon=True).start()

    def _done(self, ok, fail, is_combine=False):
        self.run_btn.configure(state="normal")
        self.update_progress(1.0)

        self.append_log(f"\n{'='*60}\n")
        self.append_log(f"COMPLETED\n")
        self.append_log(f"{'='*60}\n")
        self.append_log(f"✓ Success: {ok}\n")
        self.append_log(f"✗ Failed: {fail}\n")
        self.append_log(f"{'='*60}\n\n")

        is_po_mode = self.doc_mode.get() == "発注書→CSV"

        if fail == 0:
            if is_po_mode:
                out_txt = "1つのCSVに統合" if is_combine else f"{ok}件のCSV"
                self.update_status(f"● 完了: {out_txt} を作成", self.SUCCESS)
                messagebox.showinfo(
                    "完了",
                    f"CSVを作成しました（{ok}ファイル処理）。\n\n"
                    "⚠ ファイル名に「要確認」が付いたものは、\n"
                    "　取込前に内容をご確認ください。"
                )
            elif is_combine:
                self.update_status(f"● 完了: {ok}件を1つのExcelに統合", self.SUCCESS)
                messagebox.showinfo("Success", f"Combined {ok} PDF(s) into ONE Excel file!")
            else:
                self.update_status(f"● 完了: {ok} ファイル", self.SUCCESS)
                messagebox.showinfo("Success", f"Converted {ok} file(s) successfully!")
        else:
            if is_po_mode:
                self.update_status(f"● エラーあり (成功 {ok} / 失敗 {fail})", self.AMBER)
                messagebox.showwarning("一部エラー",
                                       f"成功: {ok}\n失敗: {fail}\n\nログをご確認ください。")
            else:
                self.update_status(f"● エラーあり (成功 {ok} / 失敗 {fail})", self.AMBER)
                messagebox.showwarning("Partial Success",
                                      f"Success: {ok}\nFailed: {fail}")


if __name__ == "__main__":
    app = InvoiceApp()
    app.mainloop()