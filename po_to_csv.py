# -*- coding: utf-8 -*-
"""
po_to_csv.py — 発注書(PO) PDF → CSV (Kintone 受注インポート用)
Pipeline: Azure DI prebuilt-layout (markdown/tables) -> deterministic parser (per vendor) -> CSV
New vendors are added via the VENDORS config only (no change to core logic).

Pure-logic module (independently testable), same style as barcode_match.py.
"""
import re
import csv
import os
from datetime import datetime

from openpyxl import Workbook

from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.ai.documentintelligence.models import (
    AnalyzeDocumentRequest,
    DocumentContentFormat,
)

# คอลัมน์ output ตาม template บริษัท (ลำดับสำคัญ)
CSV_COLUMNS = [
    "顧客", "客先注文番号", "納期",
    "製品登録/製品番号", "製品登録/製品名",
    "製品登録/注文番号", "製品登録/数量", "製品登録/単価",
]

PRODUCT_WORDS = {
    "COVER", "BRACKET", "PLATE", "GUARD", "GURAD", "DUCT", "FRAME",
    "PIPE", "BRKT", "STAY", "PLATE.", "PANEL", "ARM",
}


# ─────────────────────────────────────────────────────────────
# helpers (format ต่อเจ้า)
# ─────────────────────────────────────────────────────────────
def _clean_num(s):
    return re.sub(r"[,\s]", "", s or "")


def komatsu_dash(p):
    """品番 รูปแบบ Komatsu 3-2-5: 20Y6061151-01 -> 20Y-60-61151-01"""
    p = (p or "").strip().replace(" ", "")
    m = re.match(r"^([0-9A-Z]{3})([0-9A-Z]{2})([0-9A-Z]{5})(.*)$", p)
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}{m.group(4)}" if m else p


def product_name_only(s):
    """製品名: เอาแค่ชื่อ — ตัด tag NK(C)/XC(D) และ part code (xxx-xx-xxxxx) ออก"""
    s = re.sub(r"[A-Z]{2}\s*\([A-Z0-9]\)", " ", s or "")           # NK (C)
    s = re.sub(r"[0-9A-Z]{3}-[0-9A-Z]{2}-[0-9A-Z]{5}\S*", " ", s)  # dashed code
    toks = [t for t in s.split() if t]
    for t in toks:
        if t.upper() in PRODUCT_WORDS:
            return t
    return toks[-1] if toks else ""


# ─────────────────────────────────────────────────────────────
# VENDOR CONFIG  (เพิ่มเจ้าใหม่ที่นี่)
#   cols    : index คอลัมน์ในตาราง Azure (0-based)
#   part_fn : ฟังก์ชันแปลง 製品番号 (None = ใช้ค่าดิบ)
#   name_fn : ฟังก์ชันแปลง 製品名 (None = ใช้ค่าดิบ)
#   kokyaku : ค่า 顧客 ตายตัว (None = ดึงจากเอกสาร)
#
# NOTE: this is a demo config with a fictional sample vendor. In production,
# each real vendor gets its own entry here — adding a vendor is config-only,
# no change to the parsing logic below.
# ─────────────────────────────────────────────────────────────
VENDORS = {
    "Sample Vendor Co., Ltd. (Demo)": {
        "label": "Sample Vendor Co., Ltd. / Demo Customer Inc.",
        "cols": {"order": 0, "hinban": 1, "hinmei": 2, "qty": 3,
                 "nouki": 4, "tanka": 5, "biko": 6},
        "order_re": r"(\d{6})",             # order line number pattern (per-vendor)
        "part_fn": None,                    # 製品番号 = 品番 ดิบตรงต้นฉบับ ไม่ปรับแต่ง (ไม่ใส่ขีด)
        "name_fn": product_name_only,
        "kokyaku": None,                    # ดึงจากเอกสาร (ผู้ออก PO)
        "kokyaku_fallback": "Demo Customer Inc.",
        "kokyaku_exclude": ["Sample Vendor", "有限会社"],   # ชื่อ 発注先 (ไม่ใช่ 顧客)
        # ชื่อที่ OCR ได้ไม่ตรงกับชื่อใน ERP -> import แล้วกลายเป็นบริษัทใหม่
        # ถ้าเจอทุก keyword ในบรรทัด ให้แทนด้วยชื่อมาตรฐานที่ ERP รู้จัก
        "kokyaku_normalize": [
            (["Demo Customer", "Osaka"], "Demo Customer Inc.　Osaka Center"),
        ],
        "has_total": True,                  # ใบนี้ต้องมี 税抜金額 เสมอ -> ถ้าหายแปลว่าได้ไม่ครบหน้า
    },
}


# ─────────────────────────────────────────────────────────────
# Azure DI
# ─────────────────────────────────────────────────────────────
def analyze_layout_markdown(pdf_path, endpoint, key, log_fn=None):
    """เรียก Azure DI prebuilt-layout คืน markdown content"""
    if not endpoint or not key:
        raise RuntimeError(
            "Azure の接続情報が読み込めません。"
            "exe と同じフォルダに .env (AZURE_DOC_ENDPOINT / AZURE_DOC_KEY) を置いてください。"
        )
    client = DocumentIntelligenceClient(endpoint=endpoint,
                                        credential=AzureKeyCredential(key))
    with open(pdf_path, "rb") as f:
        data = f.read()
    poller = client.begin_analyze_document(
        "prebuilt-layout",
        AnalyzeDocumentRequest(bytes_source=data),
        output_content_format=DocumentContentFormat.MARKDOWN,
    )
    result = poller.result()
    return result.content or ""


# ─────────────────────────────────────────────────────────────
# parse markdown tables (rowspan-aware)
# ─────────────────────────────────────────────────────────────
def _parse_table_rowspan(table_html):
    """expand rowspan -> list ของ dict {col_index: value}"""
    trs = re.findall(r"<tr>(.*?)</tr>", table_html, re.S)
    grid, pending = [], {}
    for tr in trs:
        cells = re.findall(r"<t[dh]([^>]*)>(.*?)</t[dh]>", tr, re.S)
        row, c, ci = {}, 0, 0
        while ci < len(cells) or pending:
            if c in pending:
                row[c] = pending[c][0]
                pending[c][1] -= 1
                if pending[c][1] <= 0:
                    del pending[c]
                c += 1
                continue
            if ci >= len(cells):
                break
            attr, val = cells[ci]
            ci += 1
            val = re.sub(r"<[^>]+>", "", val).strip()
            rs = re.search(r'rowspan="(\d+)"', attr)
            row[c] = val
            if rs and int(rs.group(1)) > 1:
                pending[c] = [val, int(rs.group(1)) - 1]
            c += 1
        grid.append(row)
    return grid


def _merge_into(g, row, cfg):
    cc = cfg["cols"]
    for k in (cc["hinban"], cc["hinmei"], cc["nouki"]):   # text columns ต่อแถวได้
        v = row.get(k, "")
        if v and v not in (g.get(k, "") or ""):
            g[k] = (g[k] + " " + v).strip() if g.get(k) else v
    for k in (cc["qty"], cc["tanka"], cc["biko"]):        # เอาค่าแรกที่ไม่ว่าง
        if not g.get(k) and row.get(k):
            g[k] = row[k]


def parse_po_markdown(md, vendor_key):
    """คืน list ของ dict (CSV rows) + warnings list"""
    cfg = VENDORS[vendor_key]
    cc = cfg["cols"]
    order_re = cfg["order_re"]
    year = _extract_year(md)
    kokyaku = _extract_kokyaku(md, cfg)

    rows_out, warnings = [], []
    for table_html in re.findall(r"<table>(.*?)</table>", md, re.S):
        grid = _parse_table_rowspan(table_html)
        groups, cur = [], None
        for row in grid:
            col0 = row.get(cc["order"], "")
            m = re.search(order_re, col0)
            if m:
                is_hdr = ("品" in row.get(cc["hinban"], "")
                          or "数" in row.get(cc["qty"], ""))
                if cur and cur["_ord"] == m.group(1):
                    if not is_hdr:
                        _merge_into(cur, row, cfg)
                else:
                    cur = {"_ord": m.group(1)}
                    groups.append(cur)
                    if not is_hdr:
                        _merge_into(cur, row, cfg)
            else:
                joined = " ".join(str(v) for v in row.values())
                if "注文番号" in joined or "品 番" in joined or "数 量" in joined:
                    continue
                if cur:
                    _merge_into(cur, row, cfg)

        for g in groups:
            nd = re.search(r"(\d{1,2})\s*/\s*(\d{1,2})", g.get(cc["nouki"], "") or "")
            nouki = (f"{year}-{int(nd.group(1)):02d}-{int(nd.group(2)):02d}"
                     if nd and year else "")
            hinban = g.get(cc["hinban"], "")
            hinmei = g.get(cc["hinmei"], "")
            qty = _clean_num(g.get(cc["qty"], ""))
            tanka = _clean_num(g.get(cc["tanka"], ""))
            seihin_no = cfg["part_fn"](hinban) if cfg.get("part_fn") else hinban
            seihin_mei = cfg["name_fn"](hinmei) if cfg.get("name_fn") else hinmei

            row_out = {
                "顧客": kokyaku,
                "客先注文番号": g["_ord"],
                "納期": nouki,
                "製品登録/製品番号": seihin_no,
                "製品登録/製品名": seihin_mei,
                "製品登録/注文番号": g["_ord"],
                "製品登録/数量": qty,
                "製品登録/単価": tanka,
            }
            # validation ระดับแถว — ไม่เติม default ถ้าขาด, ติดธงให้คนตรวจ
            miss = [k for k in ("製品登録/数量", "製品登録/単価", "納期",
                                "製品登録/製品番号") if not row_out[k]]
            if miss:
                warnings.append(f"  ⚠ 注文番号 {g['_ord']}: {', '.join(miss)} が空欄 — 要確認")
            rows_out.append(row_out)

    return rows_out, kokyaku, year, warnings


# ─────────────────────────────────────────────────────────────
# ดึง metadata จากเอกสาร
# ─────────────────────────────────────────────────────────────
def _extract_year(md):
    m = re.search(r"(\d{4})\s*年", md)
    return m.group(1) if m else str(datetime.now().year)


def _extract_doc_total(md):
    """税抜金額 (ยอดรวมไม่รวมภาษี) สำหรับ checksum — คืน int หรือ None"""
    m = re.search(r"税抜金額\s*[:：]?\s*([\d,]+)", md)
    return int(m.group(1).replace(",", "")) if m else None


def _normalize_kokyaku(name, cfg):
    """แปลงชื่อ 顧客 ที่ OCR ได้ ให้ตรงกับชื่อใน ERP ปลายทาง (กัน import แล้วสร้างบริษัทใหม่)"""
    for keywords, canonical in cfg.get("kokyaku_normalize", []):
        if all(k in name for k in keywords):
            return canonical
    return name


def _extract_kokyaku(md, cfg):
    """顧客 = ผู้ออก PO (ดึงบรรทัดที่มี 株式会社 แต่ไม่ใช่ 発注先)"""
    if cfg.get("kokyaku"):
        return cfg["kokyaku"]
    excl = cfg.get("kokyaku_exclude", [])
    raw = cfg.get("kokyaku_fallback", "")
    for line in md.splitlines():
        line = line.strip()
        if "株式会社" in line and not any(x in line for x in excl):
            raw = line
            break
    return _normalize_kokyaku(raw, cfg)


# ─────────────────────────────────────────────────────────────
# entry: process ทีละไฟล์ + เขียน CSV + checksum
# ─────────────────────────────────────────────────────────────
def extract_po(pdf_path, vendor_key, endpoint, key, log_fn):
    """OCR + parse + checksum (ไม่เขียนไฟล์). คืน (rows, checksum_ok)
       checksum_ok: True=ตรง / False=ต้องตรวจ / None=ไม่มียอดรวมให้ตรวจ"""
    name = os.path.basename(pdf_path)
    log_fn(f"\n▶ {name}\n")
    md = analyze_layout_markdown(pdf_path, endpoint, key, log_fn)
    if not md.strip():
        raise RuntimeError("Azure DI returned empty content (OCR failed)")

    rows, kokyaku, year, warnings = parse_po_markdown(md, vendor_key)
    if not rows:
        raise RuntimeError("明細を抽出できませんでした (この仕入先のテーブル形式と不一致?)")

    log_fn(f"  顧客={kokyaku}  年度={year}  明細={len(rows)}件\n")

    # checksum: sum(数量×単価) vs 税抜金額
    doc_total = _extract_doc_total(md)
    calc = 0.0
    for r in rows:
        try:
            calc += float(r["製品登録/数量"]) * float(r["製品登録/単価"])
        except Exception:
            pass
    cfg = VENDORS[vendor_key]
    checksum_ok = None
    if doc_total is not None:
        checksum_ok = (round(calc) == doc_total)
        mark = "✅ 一致" if checksum_ok else "❌ 不一致 — 要確認"
        log_fn(f"  照合: 合計={calc:,.0f} / 伝票税抜={doc_total:,.0f}  {mark}\n")
    elif cfg.get("has_total"):
        # この伝票は必ず合計あり — 無い=全ページ未取得 (Azure tier=S0か確認)
        checksum_ok = False
        log_fn("  ❌ 税抜金額が見つかりません — 全ページ未取得の可能性 (Azure tier=S0か確認) → 要確認\n")
    else:
        log_fn("  照合: 伝票に合計なし — スキップ\n")

    for w in warnings:
        log_fn(w + "\n")
    return rows, checksum_ok


def _write_csv(rows, csv_path):
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        w.writeheader()
        w.writerows(rows)


def _write_xlsx(rows, xlsx_path):
    """xlsx — ทุกคอลัมน์เป็น Text (@) เพื่อกัน Excel แปลง type
       (เลขล้วน→number, 0 นำหน้าหาย, scientific, วันที่). ค่าคงดิบตรงต้นฉบับ.
       (ถ้าภายหลังอยากให้ 数量/単価 เป็นตัวเลขจริงเพื่อคำนวณใน Excel ค่อยแยกคอลัมน์)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "受注"
    ws.append(CSV_COLUMNS)
    for r in rows:
        ws.append([str(r.get(c, "")) for c in CSV_COLUMNS])
    for ci in range(1, len(CSV_COLUMNS) + 1):
        for ri in range(1, ws.max_row + 1):
            ws.cell(row=ri, column=ci).number_format = "@"   # Text
    wb.save(xlsx_path)


def _write_outputs(rows, out_dir, stem, fmt="both"):
    """เขียนไฟล์ output ตาม fmt: 'csv' | 'xlsx' | 'both'. คืน list ของชื่อไฟล์"""
    names = []
    if fmt in ("csv", "both"):
        p = os.path.join(out_dir, stem + ".csv")
        _write_csv(rows, p)
        names.append(os.path.basename(p))
    if fmt in ("xlsx", "both"):
        p = os.path.join(out_dir, stem + ".xlsx")
        _write_xlsx(rows, p)
        names.append(os.path.basename(p))
    return names


def export_po_to_csv(pdf_paths, out_dir, vendor_key, endpoint, key,
                     log_fn, progress_callback=None, combine=False, fmt="both"):
    """combine=False : output ต่อไฟล์
       combine=True  : รวมทุกไฟล์เป็น output เดียว (สำหรับ PO ไฟล์ละ 1 หน้า/1 ใบ)
       fmt           : 'csv' | 'xlsx' | 'both'
       คืน (ok, fail)"""
    head = "1つのCSVに統合" if combine else "ファイルごとにCSV"
    log_fn(f"\n{'='*60}\n発注書→CSV [{head}]  {vendor_key}  ({len(pdf_paths)}ファイル)\n{'='*60}\n")
    ok = fail = 0
    total = len(pdf_paths)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    if combine:
        all_rows, flagged = [], []
        for i, pdf in enumerate(pdf_paths):
            try:
                rows, cok = extract_po(pdf, vendor_key, endpoint, key, log_fn)
                all_rows.extend(rows)
                if cok is False:
                    flagged.append(os.path.basename(pdf))
                ok += 1
            except Exception as e:
                log_fn(f"  ✗ エラー: {e}\n")
                fail += 1
            if progress_callback:
                progress_callback(i + 1, total)

        if all_rows:
            suffix = "_要確認" if flagged else ""
            names = _write_outputs(all_rows, out_dir, f"統合_受注{suffix}_{ts}", fmt)
            log_fn(f"\n→ {ok}ファイル統合 = {len(all_rows)}行 → {' / '.join(names)}\n")
            if flagged:
                log_fn(f"  ⚠ 要確認ファイル ({len(flagged)}): {', '.join(flagged)}\n")
        else:
            log_fn("\n✗ 統合するデータがありません\n")
        return ok, fail

    # per-file
    for i, pdf in enumerate(pdf_paths):
        try:
            rows, cok = extract_po(pdf, vendor_key, endpoint, key, log_fn)
            base = os.path.splitext(os.path.basename(pdf))[0]
            suffix = "" if cok in (True, None) else "_要確認"
            names = _write_outputs(rows, out_dir, f"{base}_受注{suffix}_{ts}", fmt)
            log_fn(f"  → {' / '.join(names)}\n")
            ok += 1
        except Exception as e:
            log_fn(f"  ✗ エラー: {e}\n")
            fail += 1
        if progress_callback:
            progress_callback(i + 1, total)
    return ok, fail
